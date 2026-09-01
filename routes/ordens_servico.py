"""Ordem de Serviço — o documento do atendimento, próprio do site.

Decisão de 2026-08-21: não depende do AgoraOS. A OS é o registro persistente
(cliente, equipamento, defeito, status); QUEM vai atender e QUANDO continuam
sendo respondidos pelo sistema de fichas/técnicos que já existe — a OS só se
liga a um `servico` (routes/servicos.py) em vez de duplicar agenda. Uma OS
pode ter mais de uma visita ligada a ela com o tempo (voltou pra buscar
peça), por isso o vínculo mora em servicos.ordem_servico_id.
"""
import io
import json
import secrets
from datetime import datetime, timedelta, timezone

from flask import Blueprint, jsonify, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from database import db_conn, execute, fetch_all, fetch_one, insert_returning_id
from routes.clientes import criar_cliente
from routes.estoque import _norm_codigo, dar_saida
from routes.fichas import obter_ou_criar_ficha, recalcular_rota
from routes.servicos import _validar_setor
from services.fotos_extra import (adicionar_foto_extra, listar_fotos_extra,
                                  remover_foto_extra)
from services.geo import geocode_cep

ordens_servico_bp = Blueprint("ordens_servico", __name__)

# Lista fechada — mesma razão de sempre: status alimenta filtro e contagem.
STATUS_OS = [
    "aguardando_agendamento", "agendada", "em_atendimento",
    "aguardando_peca", "aguardando_orcamento", "aguardando_aprovacao",
    "aprovada", "finalizada", "cancelada",
]

TERMOS_PADRAO = (
    "O cliente autoriza a avaliação técnica do equipamento acima descrito. "
    "A taxa de avaliação é devida independentemente da aprovação do orçamento, "
    "salvo acordo em contrário. Peças e serviços só são executados mediante "
    "aprovação prévia do orçamento. A Porto Tec não se responsabiliza por "
    "defeitos pré-existentes não relacionados ao serviço solicitado, nem por "
    "dados armazenados no equipamento."
)

# Tipo de OS: pedido de 2026-08-26. Cada tipo imprime um termo diferente —
# a mesma folha de OS serve pra garantia, venda, retirada, cancelamento etc.,
# e cada situação tem uma responsabilidade jurídica distinta. Lista fechada
# pelo mesmo motivo de sempre (status, setor): dropdown, não campo livre.
TIPOS_OS = [
    "garantia_3_meses",
    "entrada_oficina",
    "saida_oficina",
    "garantia_6_meses",
    "garantia_1_ano",
    "retirada_pre_aprovada",
    "vendas",
    "retirada_aprovada",
    "retirada_orcamento",
    "acionamento_garantia_interno",
    "acionamento_garantia_externo",
    "avaliacao_tecnica",
    "cancelamento",
    "pagamento_faturamento",
    "higienizacao",
    "retirado_aprovado",
    "criterio_orcamento_reparo",
    "criterios_condicoes_orcamento",
]

TIPOS_OS_ROTULO = {
    "garantia_3_meses": "Garantia 3 meses",
    "entrada_oficina": "OS de entrada na oficina",
    "saida_oficina": "OS de saída da oficina",
    "garantia_6_meses": "OS garantia 6 meses",
    "garantia_1_ano": "OS garantia 1 ano",
    "retirada_pre_aprovada": "OS de retirada pré-aprovada",
    "vendas": "OS de vendas",
    "retirada_aprovada": "OS retirada aprovada",
    "retirada_orcamento": "OS de retirada para orçamento",
    "acionamento_garantia_interno": "Acionamento de garantia interno",
    "acionamento_garantia_externo": "Acionamento de garantia externo",
    "avaliacao_tecnica": "Avaliação técnica",
    "cancelamento": "Cancelamento",
    "pagamento_faturamento": "Pagamento / Faturamento",
    "higienizacao": "Higienização",
    "retirado_aprovado": "Retirado / Aprovado",
    "criterio_orcamento_reparo": "Critério de orçamento - reparo",
    "criterios_condicoes_orcamento": "Critérios e condições do orçamento",
}

# Termo específico de cada tipo — texto definitivo, entregue pelo Kalebe em
# 2026-08-27. \n\n vira quebra de parágrafo de verdade na impressão (ver
# .termos { white-space: pre-line } em templates/os_imprimir.html).
TERMOS_POR_TIPO = {
    "garantia_3_meses": (
        "TERMO DE GARANTIA – SERVIÇOS DE ASSISTÊNCIA TÉCNICA\n\n"
        "1. Garantia Contratual\n"
        "A empresa oferece garantia de 03 (três) meses sobre os serviços prestados, "
        "contados a partir da data da conclusão do reparo, conforme previsto no "
        "Código de Defesa do Consumidor.\n\n"
        "2. Abrangência da Garantia\n"
        "A garantia é válida exclusivamente nos bairros abrangidos pela Zona Leste "
        "da cidade de São Paulo, local de atuação da empresa contratada. Chamados "
        "fora dessa região não serão atendidos sob esta garantia.\n\n"
        "3. Limitação da Garantia\n"
        "A garantia se aplica somente às peças substituídas e aos serviços "
        "especificados na ordem de serviço, não abrangendo, em hipótese alguma, "
        "defeitos não relacionados aos itens reparados ou não mencionados no "
        "documento.\n\n"
        "4. Exclusões de Garantia\n"
        "A garantia será automaticamente invalida caso seja constatado que o "
        "defeito decorre de:\n"
        "a) Acidente, ou queda do produto e ou força maior da natureza;\n"
        "b) Mau uso ou manuseio inadequado por parte do consumidor;\n"
        "c) Ligação incorreta em rede elétrica ou uso fora das especificações "
        "técnicas do fabricante;\n"
        "d) Abertura ou tentativa de reparo por terceiros;\n"
        "e) Qualquer outra situação que comprove dolo ou culpa exclusiva do "
        "consumidor, nos termos do Art. 14, §3º, II do Código de Defesa do "
        "Consumidor.\n\n"
        "5. Taxa de Deslocamento e Avaliação\n"
        "Caso seja acionada a garantia e, durante a visita técnica, seja "
        "constatado que o problema se trata de outro defeito não relacionado ao "
        "reparo anterior ou que houve uso inadequado do equipamento, será cobrada "
        "uma taxa fixa de deslocamento e avaliação no mesmo valor da taxa de "
        "visita passado no primeiro contato com a empresa, referente a este "
        "serviço, deve ser paga no ato da avaliação, mesmo que nenhum novo reparo "
        "seja realizado.\n\n"
        "6. Aceite e Reconhecimento\n"
        "O consumidor declara estar ciente e de acordo com todas as condições "
        "deste termo, incluindo as peças utilizadas e os serviços executados, "
        "conforme registrado na ordem de serviço. Declara ainda que o equipamento "
        "foi entregue em perfeito estado de conservação e funcionamento, "
        "devidamente testado na presença do cliente."
    ),
    "entrada_oficina": (
        "Termos de entrada do Serviço\n\n"
        "1. Prazo Máximo para Retirada do Produto – Lei nº 2.560/2021\n"
        "Nos termos da Lei nº 2.560/2021, de autoria da Deputada Tayla Peres (PP), "
        "que também preside o Procon Assembleia, o cliente possui o prazo máximo "
        "de 60 (sessenta) dias corridos para realizar a retirada do equipamento "
        "deixado na assistência técnica. Decorrido esse prazo sem manifestação ou "
        "retirada, o produto será considerado abandonado, podendo ser descartado, "
        "doado ou destinado à sucata, sem direito a aviso prévio ou compensação "
        "financeira.\n\n"
        "2. Taxa de Avaliação Técnica\n"
        "Será cobrada uma taxa de avaliação técnica por produto (valor informado "
        "nesta Ordem de Serviço), nas seguintes condições: quando o orçamento "
        "apresentado não for aprovado pelo cliente; quando for constatado que o "
        "produto não apresenta defeito após os testes realizados; quando o "
        "equipamento for considerado sem possibilidade de conserto, seja por "
        "falta de peças no mercado ou inviabilidade técnica para reparo. "
        "Importante: caso o orçamento seja aprovado e o serviço realizado, a taxa "
        "de avaliação técnica será automaticamente isenta.\n\n"
        "3. Prazo para Diagnóstico e Orçamento\n"
        "O prazo médio para realização de testes, diagnóstico completo e emissão "
        "de orçamento é de até 10 (dez) dias úteis, podendo variar conforme a "
        "complexidade do equipamento ou necessidade de análise técnica mais "
        "aprofundada.\n\n"
        "4. O Cliente declara ser legítimo possuidor do bem ou estar a pedido "
        "deste.\n\n"
        "5. O Cliente está ciente que o equipamento somente poderá ser retirado "
        "com a apresentação desta Ordem de Serviço ou Documento Oficial e "
        "Original com foto. O uso indevido ou perda deste documento não são de "
        "responsabilidade da Assistência Técnica."
    ),
    "saida_oficina": (
        "TERMO DE GARANTIA\n\n"
        "1. Garantia Contratual\n"
        "A empresa oferece garantia de 03 (três) meses sobre os serviços "
        "prestados, contados a partir da data da conclusão do reparo, conforme "
        "previsto no Código de Defesa do Consumidor.\n\n"
        "2. Serviços com Possível Retorno ou Garantia\n"
        "Todo e qualquer serviço realizado em nossa unidade física, caso "
        "apresente necessidade de retorno ou solicitação de garantia, deverá ser "
        "encaminhado diretamente à loja dentro do período de cobertura "
        "estabelecido, conforme legislação vigente e política interna da "
        "empresa.\n\n"
        "3. Limitação da Garantia\n"
        "A garantia se aplica somente às peças substituídas e aos serviços "
        "especificados na ordem de serviço, não abrangendo, em hipótese alguma, "
        "defeitos não relacionados aos itens reparados ou não mencionados no "
        "documento.\n\n"
        "4. Exclusões de Garantia\n"
        "A garantia será automaticamente invalida caso seja constatado que o "
        "defeito decorre de: a) Acidente ou queda do produto; b) Mau uso ou "
        "manuseio inadequado por parte do consumidor; c) Ligação incorreta em "
        "rede elétrica ou uso fora das especificações técnicas do fabricante; "
        "d) Abertura ou tentativa de reparo por terceiros; e) Qualquer outra "
        "situação que comprove dolo ou culpa exclusiva do consumidor, nos termos "
        "do Art. 14, §3º, II do Código de Defesa do Consumidor.\n\n"
        "5. Acionamento de garantia\n"
        "Caso seja acionada a garantia e, durante a avaliação, seja constatado "
        "que o problema se trata de outro defeito não relacionado ao reparo "
        "anterior ou que houve uso inadequado do equipamento, será cobrada uma "
        "taxa fixa de avaliação no valor de R$ 40,00 (quarenta reais), a ser paga "
        "no ato da retirada, mesmo que nenhum novo reparo seja realizado.\n\n"
        "6. Aceite e Reconhecimento\n"
        "O consumidor declara estar ciente e de acordo com todas as condições "
        "deste termo, incluindo as peças utilizadas e os serviços executados, "
        "conforme registrado na ordem de serviço. Declara ainda que o equipamento "
        "foi entregue em perfeito estado de conservação e funcionamento, "
        "devidamente testado na presença do cliente."
    ),
    "garantia_6_meses": (
        "TERMO DE GARANTIA – SERVIÇOS DE ASSISTÊNCIA TÉCNICA\n\n"
        "1. Garantia Contratual\n"
        "A empresa oferece garantia de 06 (seis) meses sobre os serviços "
        "prestados, contados a partir da data da conclusão do reparo, conforme "
        "previsto no Código de Defesa do Consumidor.\n\n"
        "2. Abrangência da Garantia\n"
        "A garantia é válida exclusivamente nos bairros abrangidos pela Zona "
        "Leste da cidade de São Paulo, local de atuação da empresa contratada. "
        "Chamados fora dessa região não serão atendidos sob esta garantia.\n\n"
        "3. Limitação da Garantia\n"
        "A garantia se aplica somente às peças substituídas e aos serviços "
        "especificados na ordem de serviço, não abrangendo, em hipótese alguma, "
        "defeitos não relacionados aos itens reparados ou não mencionados no "
        "documento.\n\n"
        "4. Exclusões de Garantia\n"
        "A garantia será automaticamente invalida caso seja constatado que o "
        "defeito decorre de:\n"
        "a) Acidente, queda do produto e ou força maior da natureza;\n"
        "b) Mau uso ou manuseio inadequado por parte do consumidor;\n"
        "c) Ligação incorreta em rede elétrica ou uso fora das especificações "
        "técnicas do fabricante;\n"
        "d) Abertura ou tentativa de reparo por terceiros;\n"
        "e) Qualquer outra situação que comprove dolo ou culpa exclusiva do "
        "consumidor, nos termos do Art. 14, §3º, II do Código de Defesa do "
        "Consumidor.\n\n"
        "5. Taxa de Deslocamento e Avaliação\n"
        "Caso seja acionada a garantia e, durante a visita técnica, seja "
        "constatado que o problema se trata de outro defeito não relacionado ao "
        "reparo anterior ou que houve uso inadequado do equipamento, será cobrada "
        "uma taxa fixa de deslocamento e avaliação no mesmo valor da taxa de "
        "visita passado no primeiro contato com a empresa, referente a este "
        "serviço, deve ser paga no ato da avaliação, mesmo que nenhum novo reparo "
        "seja realizado.\n\n"
        "6. Aceite e Reconhecimento\n"
        "O consumidor declara estar ciente e de acordo com todas as condições "
        "deste termo, incluindo as peças utilizadas e os serviços executados, "
        "conforme registrado na ordem de serviço. Declara ainda que o equipamento "
        "foi entregue em perfeito estado de conservação e funcionamento, "
        "devidamente testado na presença do cliente."
    ),
    "garantia_1_ano": (
        "TERMO DE GARANTIA – SERVIÇOS DE ASSISTÊNCIA TÉCNICA\n\n"
        "1. Garantia Contratual\n"
        "A empresa oferece garantia de 12 (doze) meses sobre os serviços "
        "prestados, contados a partir da data da conclusão do reparo, conforme "
        "previsto no Código de Defesa do Consumidor.\n\n"
        "2. Abrangência da Garantia\n"
        "A garantia é válida exclusivamente nos bairros abrangidos pela Zona "
        "Leste da cidade de São Paulo, local de atuação da empresa contratada. "
        "Chamados fora dessa região não serão atendidos sob esta garantia.\n\n"
        "3. Limitação da Garantia\n"
        "A garantia se aplica somente às peças substituídas e aos serviços "
        "especificados na ordem de serviço, não abrangendo, em hipótese alguma, "
        "defeitos não relacionados aos itens reparados ou não mencionados no "
        "documento.\n\n"
        "4. Exclusões de Garantia\n"
        "A garantia será automaticamente invalida caso seja constatado que o "
        "defeito decorre de:\n"
        "a) Acidente, queda do produto e ou força maior da natureza;\n"
        "b) Mau uso ou manuseio inadequado por parte do consumidor;\n"
        "c) Ligação incorreta em rede elétrica ou uso fora das especificações "
        "técnicas do fabricante;\n"
        "d) Abertura ou tentativa de reparo por terceiros;\n"
        "e) Qualquer outra situação que comprove dolo ou culpa exclusiva do "
        "consumidor, nos termos do Art. 14, §3º, II do Código de Defesa do "
        "Consumidor.\n\n"
        "5. Taxa de Deslocamento e Avaliação\n"
        "Caso seja acionada a garantia e, durante a visita técnica, seja "
        "constatado que o problema se trata de outro defeito não relacionado ao "
        "reparo anterior ou que houve uso inadequado do equipamento, será cobrada "
        "uma taxa fixa de deslocamento e avaliação no mesmo valor da taxa de "
        "visita passado no primeiro contato com a empresa, referente a este "
        "serviço, deve ser paga no ato da avaliação, mesmo que nenhum novo reparo "
        "seja realizado.\n\n"
        "6. Aceite e Reconhecimento\n"
        "O consumidor declara estar ciente e de acordo com todas as condições "
        "deste termo, incluindo as peças utilizadas e os serviços executados, "
        "conforme registrado na ordem de serviço. Declara ainda que o equipamento "
        "foi entregue em perfeito estado de conservação e funcionamento, "
        "devidamente testado na presença do cliente."
    ),
    "retirada_pre_aprovada": (
        "RETIRADA DE EQUIPAMENTO COM PRÉ-APROVAÇÃO PARA REPARO\n\n"
        "CLÁUSULA 1 – DO OBJETO\n"
        "O presente termo refere-se à retirada técnica de equipamento(s) do "
        "cliente, previamente autorizada, para fins de realização de reparos "
        "previamente combinados com base em estimativa de valor acordada e "
        "aprovadas entre as partes. Que foi previamente informado sobre a "
        "necessidade do serviço, a estimativa de valor e demais condições "
        "básicas para execução do reparo.\n\n"
        "CLÁUSULA 2 – DA ESTIMATIVA DE VALOR\n"
        "No momento da solicitação e autorização da retirada, foi acordada uma "
        "estimativa de custo para o reparo, com base em informações preliminares "
        "e sintomas apresentados pelo equipamento no local. O valor exato será "
        "confirmado após avaliação técnica detalhada.\n\n"
        "CLÁUSULA 3 – DA DESISTÊNCIA DO SERVIÇO\n"
        "Em caso de desistência do cliente após a retirada do equipamento, será "
        "cobrada uma porcentagem fixa de 30% sobre o pré orçamento passado no "
        "local, a título de ressarcimento pelos custos operacionais com "
        "deslocamento, equipe técnica e avaliação.\n\n"
        "CLÁUSULA 4 – DA RESPONSABILIDADE DA EMPRESA\n"
        "A empresa se compromete a manusear e transportar o(s) equipamento(s) "
        "com os devidos cuidados técnicos, não se responsabilizando por vícios "
        "ocultos, defeitos pré-existentes ou falhas estruturais decorrentes do "
        "estado do produto no momento da retirada.\n\n"
        "CLÁUSULA 5 – COMUNICAÇÃO\n"
        "O orçamento será enviado pelos canais de contato fornecidos, sendo "
        "responsabilidade do cliente manter sua comunicação ativa para o "
        "andamento do serviço."
    ),
    "vendas": (
        "TERMO DE GARANTIA E ORIENTAÇÕES DE USO\n\n"
        "1. Garantia Contratual\n"
        "A empresa oferece garantia contratual de 3 (três) meses para o "
        "equipamento vendido, conforme estabelecido pelo Art. 26, inciso II, do "
        "Código de Defesa do Consumidor (CDC), iniciando-se a partir da data de "
        "retirada ou entrega do produto.\n\n"
        "2. Garantia Legal\n"
        "A garantia legal de 90 (noventa) dias é válida independente da garantia "
        "contratual e deve ser exigida exclusivamente no endereço da loja ou "
        "unidade prestadora de serviços indicada na nota fiscal ou termo de "
        "venda.\n\n"
        "3. Exclusão de Garantia\n"
        "A garantia será automaticamente invalida nos seguintes casos: mau uso "
        "do equipamento, como sobrecarga, má instalação elétrica, uso inadequado "
        "ou fora das especificações de fábrica; danos causados por quedas, "
        "impactos, transporte incorreto, infiltrações, ferrugem ou exposição "
        "excessiva à umidade; manuseio ou conserto realizado por terceiros não "
        "autorizados; falta de comprovação de aquisição, por meio de nota fiscal "
        "ou termo de compra.\n\n"
        "4. Acidentes e Defeitos por Culpa do Usuário\n"
        "Nos casos em que o defeito decorrer de acidente ou uso indevido (como "
        "inclinação excessiva, tombamento, impacto, entre outros), não será "
        "concedida garantia, conforme o Art. 14, § 3º, incisos I e II, do "
        "CDC.\n\n"
        "5. Condição do Produto no Ato da Venda\n"
        "O consumidor declara que recebeu o equipamento em pleno funcionamento e "
        "em bom estado físico e estético. A condição do equipamento é registrada "
        "por fotos, com ciência do consumidor, e comprovada por teste técnico no "
        "momento da retirada.\n\n"
        "6. Transporte do Equipamento\n"
        "A empresa não se responsabiliza por danos ocasionados durante "
        "transporte feito por terceiros ou pelo próprio cliente. Danos "
        "estruturais e/ou funcionais decorrentes de transporte inadequado anulam "
        "a garantia.\n\n"
        "7. Testes de Funcionamento\n"
        "O equipamento é testado antes da retirada ou entrega. Após esse "
        "momento, quaisquer alterações devem ser comunicadas à empresa dentro do "
        "prazo de garantia para análise técnica.\n\n"
        "8. Troca ou Substituição do Produto\n"
        "Não há substituição imediata do produto. Em caso de defeito coberto "
        "pela garantia, o equipamento será recolhido para avaliação e, se "
        "necessário, encaminhado para reparo dentro de até 30 dias, conforme o "
        "CDC.\n\n"
        "9. Prazos e Condições para Atendimento em Garantia\n"
        "Todos os atendimentos em garantia deverão ser previamente agendados. A "
        "avaliação técnica poderá levar até 5 dias úteis após o recebimento do "
        "produto na assistência.\n\n"
        "10. Taxa de Deslocamento e Avaliação\n"
        "Caso seja acionada a garantia e, durante a visita técnica, seja "
        "constatado que o problema se trata de orientação de uso, ajustes ou que "
        "houve uso inadequado do equipamento, será cobrada uma taxa fixa de "
        "deslocamento e avaliação no valor de R$ 50,00, a ser paga no ato da "
        "visita, mesmo que nenhum novo reparo seja realizado.\n\n"
        "ORIENTAÇÕES PARA TRANSPORTE E USO DO ELETRODOMÉSTICO\n\n"
        "1. Não tombar, deitar ou inclinar a geladeira superior a 45°. Isso pode "
        "fazer com que os fluidos internos se movimentem de forma errada e "
        "comprometam o sistema de refrigeração.\n"
        "2. Após transporte, aguarde no mínimo 6 horas antes de ligar o "
        "equipamento na tomada, para que os fluidos retornem ao seu local de "
        "origem.\n"
        "3. Em caso de quedas de energia, desligue o equipamento imediatamente e "
        "aguarde estabilização da rede elétrica antes de religá-lo, para "
        "proteger os componentes internos.\n"
        "4. Nunca apoie ou puxe o equipamento pelos canos (tubos traseiros), "
        "pois eles são sensíveis e podem ser danificados com facilidade.\n"
        "5. Evite expor o equipamento ao sol ou à chuva, mesmo que por curtos "
        "períodos.\n"
        "6. Não use extensões ou adaptadores múltiplos para ligar o equipamento, "
        "a fim de evitar sobrecarga elétrica.\n\n"
        "Declaro que li e estou de acordo com todas as cláusulas acima."
    ),
    "retirada_aprovada": (
        "RETIRADA DE EQUIPAMENTO COM APROVAÇÃO PARA REPARO\n\n"
        "CLÁUSULA 1 – DO OBJETO\n"
        "O presente termo refere-se à retirada técnica de equipamento(s) do "
        "cliente, autorizada, para fins de realização de reparos previamente "
        "combinados com base em estimativa de valor acordada e aprovadas entre "
        "as partes.\n\n"
        "CLÁUSULA 2 – DA APROVAÇÃO PELO CLIENTE\n"
        "A retirada foi realizada mediante autorização expressa do cliente, que "
        "foi previamente informado sobre a necessidade do serviço.\n\n"
        "CLÁUSULA 3 – DA ESTIMATIVA DE VALOR\n"
        "No momento da solicitação e autorização da retirada, foi acordada uma "
        "estimativa de custo para o reparo, com base em informações preliminares "
        "e sintomas apresentados pelo equipamento no local.\n\n"
        "CLÁUSULA 4 – DA DESISTÊNCIA DO SERVIÇO\n"
        "Em caso de desistência do cliente após a retirada do equipamento, será "
        "cobrada uma porcentagem fixa de 30% sobre o pré orçamento passado no "
        "local, a título de ressarcimento pelos custos operacionais com "
        "deslocamento, equipe técnica e avaliação.\n\n"
        "CLÁUSULA 5 – DA RESPONSABILIDADE DA EMPRESA\n"
        "A empresa se compromete a manusear e transportar o(s) equipamento(s) "
        "com os devidos cuidados técnicos, não se responsabilizando por vícios "
        "ocultos, defeitos pré-existentes ou falhas estruturais decorrentes do "
        "estado do produto no momento da retirada.\n\n"
        "CLÁUSULA 6 – COMUNICAÇÃO\n"
        "O orçamento será enviado pelos canais de contato fornecidos, sendo "
        "responsabilidade do cliente manter sua comunicação ativa para o "
        "andamento do serviço."
    ),
    "retirada_orcamento": (
        "TERMO DE RETIRADA PARA AVALIAÇÃO TÉCNICA EM OFICINA\n\n"
        "1. Finalidade da Retirada\n"
        "O equipamento descrito abaixo foi recolhido exclusivamente para análise "
        "técnica, testes em bancada e elaboração de orçamento detalhado. A "
        "retirada não implica autorização prévia de conserto, sendo necessária "
        "aprovação formal do cliente após o envio do diagnóstico.\n\n"
        "2. Prazos e Condições\n"
        "O prazo médio para diagnóstico e orçamento é de até 5 dias úteis, "
        "podendo variar conforme complexidade e disponibilidade de peças. Caso o "
        "cliente não aprove o orçamento em até 10 dias corridos após envio, ou "
        "não responda ou não retorne nossas tentativas de contato, o "
        "equipamento será mantido em depósito por até 30 dias, e após esse "
        "período, poderá ser considerado abandonado, conforme art. 1.275, III "
        "do Código Civil. O transporte e a guarda serão realizados pela "
        "Portotec com todo o zelo necessário, não nos responsabilizando por "
        "defeitos ocultos, oxidação, trincas pré-existentes ou danos internos "
        "decorrentes de mau uso ou desgaste natural.\n\n"
        "3. Custos de Avaliação\n"
        "Caso o cliente opte por não realizar o reparo, poderá ser cobrada taxa "
        "de diagnóstico (R$ 30,00) e deslocamento, conforme quilometragem "
        "rodada, informaremos o cliente com base nisso.\n\n"
        "4. Aprovação do Cliente\n"
        "O orçamento será enviado por WhatsApp, devendo o cliente confirmar "
        "expressamente a autorização para o conserto e valores antes do início "
        "de qualquer intervenção.\n\n"
        "5. Retorno do Equipamento\n"
        "Após os testes e/ou orçamento, o equipamento será devolvido nas mesmas "
        "condições em que foi recolhido, salvo autorização expressa para "
        "reparo, uma vez que temos registro de como foi retirado e como estamos "
        "entregando.\n\n"
        "6. Declaração do Cliente\n"
        "Declaro que li e estou ciente dos termos acima, autorizando a retirada "
        "do equipamento para diagnóstico e orçamento, e que a Portotec não se "
        "responsabiliza por eventuais defeitos que não estejam relacionados ao "
        "processo de teste ou transporte devidamente comprovado."
    ),
    "acionamento_garantia_interno": (
        "TERMO DE ACIONAMENTO DE GARANTIA\n"
        "Produtos de pequeno porte e analisados em oficina\n\n"
        "1. Natureza do Atendimento\n"
        "Este atendimento refere-se a um retorno em garantia, dentro do prazo "
        "legal de 03 (três) meses, contado da data do reparo anterior, conforme "
        "Código de Defesa do Consumidor e Termo de Garantia fornecido pela "
        "empresa.\n\n"
        "2. Avaliação do Produto\n"
        "O equipamento será submetido à análise técnica para verificação da "
        "causa do defeito relatado. Caso seja confirmado que o problema está "
        "relacionado ao reparo anterior, a correção será realizada sem qualquer "
        "custo adicional ao cliente. Caso seja constatado que o defeito decorre "
        "de mau uso, acidente, manuseio inadequado, uso fora das especificações "
        "do fabricante, abertura por terceiros ou qualquer situação não "
        "relacionada ao reparo anterior, a garantia não será aplicada.\n\n"
        "3. Taxa de Avaliação\n"
        "Se, durante a análise, for identificado que o defeito não possui "
        "relação com o reparo anterior, será cobrada uma taxa de avaliação no "
        "valor de R$ 40,00 (quarenta reais), a ser paga no ato da retirada do "
        "equipamento, mesmo que o cliente não opte por realizar um novo "
        "reparo.\n\n"
        "4. Responsabilidade do Cliente\n"
        "O cliente declara estar ciente de que este termo não amplia nem altera "
        "as condições de garantia previamente fornecidas, limitando-se apenas ao "
        "atendimento referente à Ordem de Serviço mencionada.\n\n"
        "5. Aceite e Assinatura\n"
        "Ao assinar este termo, o cliente reconhece ter ciência de todas as "
        "condições descritas, autorizando a análise e estando ciente da "
        "possibilidade de cobrança da taxa de avaliação, caso aplicável."
    ),
    "acionamento_garantia_externo": (
        "TERMO DE ACIONAMENTO DE GARANTIA\n\n"
        "1. Natureza do Atendimento\n"
        "Este atendimento refere-se a um retorno em garantia, dentro do prazo "
        "legal de 03 (três) meses, contado da data do reparo anterior, conforme "
        "Código de Defesa do Consumidor e Termo de Garantia fornecido pela "
        "empresa.\n\n"
        "2. Avaliação do Produto\n"
        "O equipamento será submetido à análise técnica para verificação da "
        "causa do defeito relatado. Caso seja confirmado que o problema está "
        "relacionado ao reparo anterior, a correção será realizada sem qualquer "
        "custo adicional ao cliente. Caso seja constatado que o defeito decorre "
        "de mau uso, acidente, manuseio inadequado, uso fora das especificações "
        "do fabricante, abertura por terceiros ou qualquer situação não "
        "relacionada ao reparo anterior, a garantia não será aplicada.\n\n"
        "3. Taxa de Avaliação\n"
        "Se, durante a análise, for identificado que o defeito não possui "
        "relação com o reparo anterior, será cobrada uma taxa de avaliação no "
        "valor de R$ 40,00 (quarenta reais), a ser paga no ato da retirada do "
        "equipamento, mesmo que o cliente não opte por realizar um novo "
        "reparo.\n\n"
        "4. Responsabilidade do Cliente\n"
        "O cliente declara estar ciente de que este termo não amplia nem altera "
        "as condições de garantia previamente fornecidas, limitando-se apenas ao "
        "atendimento referente à Ordem de Serviço mencionada.\n\n"
        "5. Aceite e Assinatura\n"
        "Ao assinar este termo, o cliente reconhece ter ciência de todas as "
        "condições descritas, autorizando a análise e estando ciente da "
        "possibilidade de cobrança da taxa de avaliação, caso aplicável."
    ),
    "avaliacao_tecnica": (
        "Taxa de avaliação\n\n"
        "O valor pago é o valor acordado via WhatsApp e nunca um valor "
        "diferente; em casos de valores diferentes, pode e deve ser contactada "
        "a CONTRATADA via WhatsApp para verificação de veracidade.\n\n"
        "Pagamento da taxa deve ser feito no ATO DO ATENDIMENTO; em caso de "
        "falta deste pagamento o cliente fica sujeito a protesto em seu nome no "
        "SERASA.\n\n"
        "A taxa é cobrada apenas se: o orçamento não for aprovado; o "
        "equipamento não tiver conserto (por ausência de peças no mercado ou "
        "impossibilidade técnica de reparo); o atendimento for apenas para "
        "orientações de uso; ou, em casos diferentes, se a cotação de peça "
        "precisar ser feita fora do local de atendimento com consulta a "
        "fabricante ou fornecedores externos.\n\n"
        "Formas de pagamento: Pix, dinheiro ou cartão."
    ),
    "cancelamento": (
        "TERMO DE CANCELAMENTO DE SERVIÇO\n\n"
        "Declaramos, para os devidos fins, que o serviço contratado foi "
        "cancelado de comum acordo entre as partes. As peças foram devidamente "
        "retiradas, não havendo quaisquer pendências entre o Cliente e a "
        "Contratada.\n\n"
        "Esclarece-se que o pagamento foi devidamente devolvido ao cliente, "
        "pelo mesmo meio.\n\n"
        "Este termo é firmado em comum acordo, estando ambas as partes cientes "
        "e de pleno consentimento."
    ),
    "pagamento_faturamento": (
        "TERMO DE PAGAMENTO FATURADO\n\n"
        "CLÁUSULA 1 – DO OBJETO\n"
        "1.1 A CONTRATADA prestará serviços de assistência técnica, manutenção, "
        "reparo, instalação e demais atividades correlatas em eletrodomésticos "
        "descritos em cada Ordem de Serviço (OS) emitida.\n\n"
        "CLÁUSULA 2 – DO PAGAMENTO\n"
        "2.1 O pagamento será efetuado de forma faturada, conforme prazo "
        "acordado previamente entre as partes, podendo ser 15, 20 ou 30 dias "
        "após a emissão da Nota Fiscal.\n"
        "2.2 O não pagamento no prazo estabelecido acarretará multa de 2% sobre "
        "o valor total da fatura, juros de 1% ao mês e correção monetária com "
        "base no índice aplicável ao período.\n"
        "2.3 A ausência de pagamento poderá resultar em suspensão imediata dos "
        "atendimentos até a regularização financeira.\n\n"
        "CLÁUSULA 3 – DAS RESPONSABILIDADES DA CONTRATADA\n"
        "3.1 Realizar os serviços dentro de padrões técnicos adequados, "
        "utilizando peças originais quando necessário, observando "
        "procedimentos recomendados pelos fabricantes.\n"
        "3.2 Emitir Nota Fiscal referente aos serviços realizados.\n\n"
        "CLÁUSULA 4 – DAS RESPONSABILIDADES DA CONTRATANTE\n"
        "4.1 Cumprir rigorosamente os prazos de pagamento acordados.\n"
        "4.2 Garantir a segurança e acesso ao local dos serviços.\n"
        "4.3 Prestar informações verdadeiras sobre os equipamentos atendidos.\n\n"
        "CLÁUSULA 5 – DA GARANTIA\n"
        "5.1 Os serviços possuem garantia técnica de até 90 dias, limitada "
        "exclusivamente ao serviço executado e peças substituídas (quando "
        "aplicável), conforme legislação vigente.\n"
        "5.2 A garantia não cobre danos decorrentes de mau uso, instalação "
        "incorreta por terceiros, oscilações elétricas, quedas, impactos e "
        "quaisquer fatores externos.\n\n"
        "CLÁUSULA 6 – DA INADIMPLÊNCIA\n"
        "6.1 Em caso de atraso superior a 30 dias, a CONTRATADA poderá incluir "
        "o débito em plataformas de negociação e órgãos de proteção ao crédito "
        "(SPC/Serasa), bem como protestar o título em cartório, "
        "independentemente de aviso prévio.\n"
        "6.2 Persistindo o inadimplemento, a CONTRATADA poderá acionar o "
        "departamento jurídico para cobrança judicial, sendo o CONTRATANTE "
        "responsável por honorários advocatícios de até 20% sobre o valor "
        "total do débito.\n"
        "6.3 Todas as despesas extras de cobrança serão integralmente "
        "repassadas ao CONTRATANTE.\n\n"
        "CLÁUSULA 7 – DA RESCISÃO\n"
        "7.1 O presente termo poderá ser rescindido por qualquer das partes, "
        "mediante aviso com antecedência mínima de 10 dias, sem prejuízo das "
        "obrigações pendentes.\n\n"
        "CLÁUSULA 8 – DO FORO\n"
        "8.1 Para dirimir eventuais conflitos, as partes elegem o foro da "
        "Comarca de São Paulo/SP, renunciando a qualquer outro, por mais "
        "privilegiado que seja."
    ),
    "higienizacao": (
        "TERMO DE HIGIENIZAÇÃO\n\n"
        "1. Natureza do Serviço\n"
        "O serviço realizado trata-se exclusivamente de higienização técnica e "
        "preventiva, sem substituição de peças, componentes eletrônicos ou "
        "mecânicos. Portanto, não se aplica termo de garantia, uma vez que não "
        "houve reparo, instalação ou troca de partes do produto.\n\n"
        "2. Condições do Equipamento\n"
        "Antes do serviço, o equipamento foi inspecionado e apresentava "
        "funcionamento normal, conforme informado pelo cliente. Caso haja "
        "defeito posterior relacionado a peças, desgaste natural ou mau uso, o "
        "mesmo não está vinculado à higienização realizada.\n\n"
        "3. Responsabilidade Técnica\n"
        "O serviço foi executado por profissional qualificado, seguindo os "
        "procedimentos técnicos recomendados pelo fabricante e com produtos "
        "adequados à limpeza de eletrodomésticos.\n\n"
        "4. Risco e Limitações\n"
        "Durante o processo de higienização, podem ser identificados defeitos "
        "preexistentes, oxidações internas, resíduos em locais de difícil "
        "acesso ou desgastes naturais — que não são de responsabilidade do "
        "prestador. Não nos responsabilizamos por falhas que surjam "
        "posteriormente, decorrentes de componentes antigos, infiltração, "
        "curto-circuito, sobrecarga elétrica ou mau uso.\n\n"
        "5. Prazo de Validade do Serviço\n"
        "O serviço é pontual e preventivo, sendo recomendada nova higienização "
        "a cada 6 meses (lavadoras, lava e seca, geladeiras) ou conforme uso e "
        "ambiente. É fornecido comprovante de execução, não configurando termo "
        "de garantia.\n\n"
        "6. Concordância do Cliente\n"
        "Ao assinar ou confirmar o atendimento, o cliente declara estar ciente "
        "de que se trata de serviço de manutenção preventiva, sem garantia por "
        "troca de peças, e que o equipamento encontrava-se em funcionamento no "
        "ato da visita."
    ),
    "retirado_aprovado": (
        "TERMO DE RETIRADA SEM REPARO E ENCERRAMENTO DE RESPONSABILIDADE\n\n"
        "Declaro, para os devidos fins, que estou retirando o produto nesta "
        "data, ciente de que não foi realizado reparo pela assistência técnica, "
        "conforme informado previamente.\n\n"
        "Declaro ainda que conferi o produto no ato da retirada e que ele está "
        "sendo entregue nas mesmas condições em que foi recebido pela empresa, "
        "sem qualquer avaria aparente ocasionada durante sua permanência na "
        "assistência.\n\n"
        "A partir da assinatura deste termo e da retirada do produto, declaro "
        "estar ciente de que a empresa fica isenta de qualquer responsabilidade, "
        "vínculo, guarda, garantia técnica ou obrigação futura referente ao "
        "produto, uma vez que o equipamento está sendo retirado sem a execução "
        "do reparo.\n\n"
        "Assumo total responsabilidade pelo produto a partir deste momento, "
        "nada tendo a reclamar posteriormente quanto ao seu estado, "
        "funcionamento, defeitos existentes ou eventuais problemas decorrentes "
        "da não realização do reparo."
    ),
    "criterio_orcamento_reparo": (
        "CRITÉRIOS E CONDIÇÕES DO ORÇAMENTO\n\n"
        "1. Validade do orçamento\n"
        "Este orçamento é válido por até 15 (quinze) dias, contados a partir da "
        "data de sua emissão, desde que não ocorram alterações de valores, "
        "disponibilidade de peças ou condições comerciais por parte dos "
        "fabricantes e fornecedores.\n\n"
        "2. Forma de pagamento\n"
        "Aceitamos pagamentos via PIX, dinheiro, cartão de débito ou crédito. "
        "Para parcelamentos, consulte nossa equipe para verificar as melhores "
        "condições disponíveis.\n\n"
        "3. Valor da vistoria técnica\n"
        "Caso tenha sido realizado pagamento referente à vistoria ou avaliação "
        "técnica, informamos que esse valor já estará integralmente descontado "
        "do valor total do reparo apresentado neste orçamento, não havendo "
        "cobrança em duplicidade.\n\n"
        "4. Encomenda de peças\n"
        "Quando houver necessidade de encomenda de alguma peça que não esteja "
        "disponível em nosso estoque próprio, será solicitado um sinal de "
        "R$ 50,00 para confirmação e andamento do pedido.\n\n"
        "O valor pago como sinal será integralmente descontado do valor total "
        "do reparo, no momento da finalização do serviço.\n\n"
        "5. Aprovação do orçamento\n"
        "A execução do serviço e/ou solicitação de peças será realizada após a "
        "aprovação do orçamento pelo cliente.\n\n"
        "6. Disponibilidade de peças\n"
        "Os prazos para fornecimento de peças poderão variar de acordo com a "
        "disponibilidade do fabricante, distribuidor ou fornecedor responsável.\n\n"
        "7. Alterações no orçamento\n"
        "Caso sejam identificados defeitos adicionais, peças ou serviços não "
        "previstos inicialmente, o cliente será previamente comunicado. Nenhum "
        "valor adicional será executado sem nova autorização.\n\n"
        "8. Prazo de execução\n"
        "Os prazos informados são estimados e podem sofrer alterações em razão "
        "da disponibilidade de peças, necessidade de testes, complexidade do "
        "equipamento ou fatores externos.\n\n"
        "9. Garantia\n"
        "A garantia refere-se exclusivamente às peças substituídas e aos "
        "serviços efetivamente executados, conforme condições e prazo "
        "informados na Ordem de Serviço.\n\n"
        "Ao aprovar este orçamento, o cliente declara estar ciente e de acordo "
        "com os valores, serviços e condições acima descritos."
    ),
    # Texto idêntico ao "criterio_orcamento_reparo" (17) — o Kalebe entregou
    # os dois separados em 2026-08-28, título diferente ("Critérios e
    # condições do orçamento"), mesmo conteúdo. Mantido como opção própria
    # de propósito, não uma duplicata pra remover.
    "criterios_condicoes_orcamento": (
        "CRITÉRIOS E CONDIÇÕES DO ORÇAMENTO\n\n"
        "1. Validade do orçamento\n"
        "Este orçamento é válido por até 15 (quinze) dias, contados a partir da "
        "data de sua emissão, desde que não ocorram alterações de valores, "
        "disponibilidade de peças ou condições comerciais por parte dos "
        "fabricantes e fornecedores.\n\n"
        "2. Forma de pagamento\n"
        "Aceitamos pagamentos via PIX, dinheiro, cartão de débito ou crédito. "
        "Para parcelamentos, consulte nossa equipe para verificar as melhores "
        "condições disponíveis.\n\n"
        "3. Valor da vistoria técnica\n"
        "Caso tenha sido realizado pagamento referente à vistoria ou avaliação "
        "técnica, informamos que esse valor já estará integralmente descontado "
        "do valor total do reparo apresentado neste orçamento, não havendo "
        "cobrança em duplicidade.\n\n"
        "4. Encomenda de peças\n"
        "Quando houver necessidade de encomenda de alguma peça que não esteja "
        "disponível em nosso estoque próprio, será solicitado um sinal de "
        "R$ 50,00 para confirmação e andamento do pedido.\n\n"
        "O valor pago como sinal será integralmente descontado do valor total "
        "do reparo, no momento da finalização do serviço.\n\n"
        "5. Aprovação do orçamento\n"
        "A execução do serviço e/ou solicitação de peças será realizada após a "
        "aprovação do orçamento pelo cliente.\n\n"
        "6. Disponibilidade de peças\n"
        "Os prazos para fornecimento de peças poderão variar de acordo com a "
        "disponibilidade do fabricante, distribuidor ou fornecedor responsável.\n\n"
        "7. Alterações no orçamento\n"
        "Caso sejam identificados defeitos adicionais, peças ou serviços não "
        "previstos inicialmente, o cliente será previamente comunicado. Nenhum "
        "valor adicional será executado sem nova autorização.\n\n"
        "8. Prazo de execução\n"
        "Os prazos informados são estimados e podem sofrer alterações em razão "
        "da disponibilidade de peças, necessidade de testes, complexidade do "
        "equipamento ou fatores externos.\n\n"
        "9. Garantia\n"
        "A garantia refere-se exclusivamente às peças substituídas e aos "
        "serviços efetivamente executados, conforme condições e prazo "
        "informados na Ordem de Serviço.\n\n"
        "Ao aprovar este orçamento, o cliente declara estar ciente e de acordo "
        "com os valores, serviços e condições acima descritos."
    ),
}


# Modelo do documento — pedido de 2026-08-27. A mesma tabela ordens_servico
# serve pros três, porque os três compartilham cliente/equipamento/status;
# o que muda é qual conjunto de campos extra se aplica e qual seção some ou
# aparece na impressão (ver templates/os_imprimir.html):
#   "os"              — o padrão de sempre: tipo_os + termo jurídico + taxa.
#   "chamado_tecnico" — registro pontual de atendimento: foto, defeito,
#                       solução, técnico que atendeu.
#   "orcamento"       — proposta de preço: defeito, solução, lista de itens
#                       (serviço + valor) em ordem_servico_itens.
MODELOS_OS = ["chamado_tecnico", "os", "orcamento"]

MODELOS_OS_ROTULO = {
    "os": "Ordens de Serviço",
    "chamado_tecnico": "Chamado Técnico",
    "orcamento": "Orçamento",
}

FOTO_MAXIMA = 900 * 1024
PREFIXOS_FOTO = ("data:image/jpeg;base64,", "data:image/png;base64,",
                 "data:image/webp;base64,")


def _foto_valida(foto):
    if not isinstance(foto, str) or not foto.startswith(PREFIXOS_FOTO):
        return None
    if len(foto) > FOTO_MAXIMA:
        return None
    return foto


def _validar_modelo_os(valor):
    """Devolve (mensagem_de_erro, modelo). Vazio ('' -> None) cai no padrão
    'os' — toda OS de antes desse recurso é desse modelo, e a maioria das
    novas continua sendo."""
    modelo = (valor or "os").strip()
    if modelo not in MODELOS_OS:
        return "Modelo de OS inválido.", None
    return "", modelo


def _agora() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _quem() -> str:
    return session.get("usuario_nome") or "Administrador"


def _campos_os(d: dict) -> dict:
    return {
        "tipo_aparelho": (d.get("tipo_aparelho") or "").strip(),
        "marca": (d.get("marca") or "").strip(),
        "modelo": (d.get("modelo") or "").strip(),
        "numero_serie": (d.get("numero_serie") or "").strip(),
        "voltagem": (d.get("voltagem") or "").strip(),
        "acessorios": (d.get("acessorios") or "").strip(),
        "defeito_declarado": (d.get("defeito_declarado") or "").strip(),
        "observacao": (d.get("observacao") or "").strip(),
    }


def _validar_tipo_os(valor):
    """Devolve (mensagem_de_erro, tipo). Erro vazio significa que passou —
    mesmo padrão de _validar_setor em routes/servicos.py."""
    tipo = (valor or "").strip()
    if not tipo:
        return "Escolha o tipo de OS.", None
    if tipo not in TIPOS_OS:
        return "Tipo de OS inválido.", None
    return "", tipo


def _num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# Campos/seções que dá pra esconder NA IMPRESSÃO sem apagar o dado — o
# checkbox desmarcado só tira do papel, o campo continua salvo e editável
# (pedido de 2026-08-28: "quero essa casinha em tudo", não só num grupinho
# de 5 no fim do formulário). "valores" do Orçamento fica de fora de
# propósito — esconder ali derrotaria o próprio documento.
SECOES_IMPRIMIVEIS = {
    "tipo_aparelho", "marca", "modelo", "numero_serie", "voltagem", "acessorios",
    "defeito_declarado", "solucao", "foto", "tecnico", "forma_pagamento", "taxa",
    "observacao", "valores", "garantia", "termos",
}


def _validar_imprimir_ocultar(valor) -> str:
    if not isinstance(valor, list):
        return "[]"
    return json.dumps([v for v in valor if v in SECOES_IMPRIMIVEIS])


def _validar_garantia_meses(valor):
    try:
        n = int(valor)
    except (TypeError, ValueError):
        return None
    return n if n in (3, 6, 12) else None


def _validar_data_iso(valor):
    """'2026-08-29' -> mesma string; qualquer outra coisa (vazio, formato
    errado, injeção de texto livre) vira None — melhor sem data do que uma
    data que quebra o cálculo de garantia na impressão."""
    texto = (valor or "").strip()
    if not texto:
        return None
    try:
        datetime.strptime(texto, "%Y-%m-%d")
    except ValueError:
        return None
    return texto


@ordens_servico_bp.route("/ordens-servico/metricas", methods=["GET"])
def metricas():
    """OS por mês (últimos 6), tempo médio até finalizar, e indicação que
    mais traz cliente — calculado em Python pra não depender de função de
    data que diverge entre SQLite e Postgres."""
    with db_conn() as conn:
        ordens = fetch_all(conn, "SELECT criado_em, finalizada_em, status FROM ordens_servico")
        indicacoes = fetch_all(conn, """
            SELECT c.indicacao FROM clientes c
             WHERE c.indicacao IS NOT NULL AND c.indicacao <> ''
               AND EXISTS (SELECT 1 FROM ordens_servico os WHERE os.cliente_id = c.id)
        """)

    por_mes = {}
    for o in ordens:
        mes = (o.get("criado_em") or "")[:7]  # "AAAA-MM"
        if mes:
            por_mes[mes] = por_mes.get(mes, 0) + 1
    meses_ordenados = sorted(por_mes.keys())[-6:]

    duracoes = []
    for o in ordens:
        if o.get("finalizada_em") and o.get("criado_em"):
            inicio = _parse_data_hora(o["criado_em"])
            fim = _parse_data_hora(o["finalizada_em"])
            if inicio and fim and fim >= inicio:
                duracoes.append((fim - inicio).total_seconds() / 86400)
    tempo_medio_dias = round(sum(duracoes) / len(duracoes), 1) if duracoes else None

    contagem_indicacao = {}
    for i in indicacoes:
        chave = i["indicacao"]
        contagem_indicacao[chave] = contagem_indicacao.get(chave, 0) + 1
    por_indicacao = sorted(contagem_indicacao.items(), key=lambda x: -x[1])

    return jsonify({
        "por_mes": [{"mes": m, "total": por_mes[m]} for m in meses_ordenados],
        "tempo_medio_dias": tempo_medio_dias,
        "os_finalizadas_com_tempo": len(duracoes),
        "por_indicacao": [{"indicacao": k, "total": v} for k, v in por_indicacao],
        "total_geral": len(ordens),
    })


def _parse_data_hora(texto):
    texto = str(texto or "").strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


@ordens_servico_bp.route("/ordens-servico/status", methods=["GET"])
def listar_status():
    return jsonify({"status": STATUS_OS})


@ordens_servico_bp.route("/ordens-servico/tipos", methods=["GET"])
def listar_tipos():
    return jsonify({"tipos": [{"chave": t, "rotulo": TIPOS_OS_ROTULO[t]} for t in TIPOS_OS]})


@ordens_servico_bp.route("/ordens-servico/modelos", methods=["GET"])
def listar_modelos():
    return jsonify({"modelos": [{"chave": m, "rotulo": MODELOS_OS_ROTULO[m]} for m in MODELOS_OS]})


# Tipos de aparelho comuns na loja (pedido de 2026-09-01: "microondas,
# purificador de água, air fryer") — semente pra já aparecer sugestão antes
# de qualquer OS existir com esses valores. Cresce sozinho depois: ver
# listar_tipos_aparelho, que junta isto com os valores já usados de verdade.
TIPOS_APARELHO_SEMENTE = [
    "Micro-ondas", "Purificador de Água", "Air Fryer", "Geladeira",
    "Freezer", "Lava-louças", "Lavadora", "Secadora", "Adega",
    "Cafeteira", "Liquidificador", "Ventilador", "Ar-condicionado",
    "Aspirador de Pó", "Fritadeira Elétrica", "Forno Elétrico",
]


@ordens_servico_bp.route("/tipos-aparelho", methods=["GET"])
def listar_tipos_aparelho():
    """Sugestões pro campo "Tipo" do Equipamento — mistura a semente fixa com
    o que já foi digitado de verdade em OS anteriores, sem duplicar (mesmo
    nome com capitalização diferente vira um só, prevalece a semente/o mais
    usado). Não é obrigatório escolher da lista — datalist, não select."""
    with db_conn() as conn:
        usados = fetch_all(conn, """
            SELECT tipo_aparelho, COUNT(*) AS n FROM ordens_servico
             WHERE tipo_aparelho IS NOT NULL AND tipo_aparelho != ''
             GROUP BY tipo_aparelho ORDER BY n DESC
        """)
    vistos = {}
    for nome in TIPOS_APARELHO_SEMENTE:
        vistos[nome.lower()] = nome
    for l in usados:
        chave = l["tipo_aparelho"].strip().lower()
        if chave and chave not in vistos:
            vistos[chave] = l["tipo_aparelho"].strip()
    return jsonify({"tipos": sorted(vistos.values(), key=str.lower)})


@ordens_servico_bp.route("/catalogo-servicos", methods=["GET"])
def listar_catalogo_servicos():
    """Lista de serviço+valor já cadastrados antes, pra reaproveitar num
    orçamento novo sem redigitar — ordenado por uso mais recente primeiro."""
    with db_conn() as conn:
        itens = fetch_all(conn, "SELECT * FROM catalogo_servicos ORDER BY id DESC")
    return jsonify({"itens": itens})


@ordens_servico_bp.route("/catalogo-servicos", methods=["POST"])
def criar_catalogo_servico():
    d = request.get_json(silent=True) or {}
    nome = (d.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "Informe o nome do serviço"}), 400
    valor = _num(d.get("valor"))

    with db_conn(commit=True) as conn:
        existente = fetch_one(conn, "SELECT id FROM catalogo_servicos WHERE LOWER(nome) = LOWER(?)",
                              (nome,))
        if existente:
            # Reaproveita em vez de duplicar: dois itens "Troca de resistência"
            # com valores diferentes na lista de sugestão só confundiria na
            # hora de escolher qual usar. Atualiza o valor pro mais recente.
            execute(conn, "UPDATE catalogo_servicos SET valor = ? WHERE id = ?",
                   (valor, existente["id"]))
            return jsonify({"mensagem": "Serviço já existia — valor atualizado",
                            "id": existente["id"]})
        item_id = insert_returning_id(conn, """
            INSERT INTO catalogo_servicos (nome, valor, criado_em, criado_por)
            VALUES (?, ?, ?, ?)
        """, (nome, valor, _agora(), _quem()))

    return jsonify({"mensagem": "Serviço adicionado ao catálogo", "id": item_id}), 201


@ordens_servico_bp.route("/catalogo-servicos/<int:item_id>", methods=["DELETE"])
def remover_catalogo_servico(item_id):
    with db_conn(commit=True) as conn:
        execute(conn, "DELETE FROM catalogo_servicos WHERE id = ?", (item_id,))
    return jsonify({"mensagem": "Removido do catálogo"})


@ordens_servico_bp.route("/ordens-servico", methods=["GET"])
def listar():
    """?status filtra; ?cliente_id filtra por cliente; ?busca acha por número
    da OS ou nome do cliente — telefone toca e alguém pergunta "cadê a OS 12",
    não dá pra obrigar a procurar folheando por status. ?dias=N filtra pelas
    abertas nos últimos N dias — mesma convenção de /desfechos e /historico.

    ?fonte=peca|reagendamento separa a fila de Agendar Clientes em dois lados,
    do jeito que o Verificador de CEP já separa em abas: 'peca' é toda OS
    referenciada por pecas_chegada.ordem_servico_id (peça chegou, aba Peças
    mandou pra cá); 'reagendamento' é o resto (OS nova nunca agendada, ou
    técnico marcou volto_depois/reagendar em campo). Sem coluna nova — o
    próprio pecas_chegada já é a marca de qual lado é qual.

    Filhas (Chamado Técnico/Orçamento pendurado numa OS já existente — ver
    os_pai_id) NÃO aparecem aqui por padrão, pra lista principal não
    empilhar um monte de linha do mesmo caso — só aparecem com
    ?pai_id=<id>, que devolve exatamente as filhas daquele pai.

    ?origem=panasonic|nossa separa a aba OS em dois lados — pedido de
    2026-08-28: "OS Panasonic" é toda OS que nasceu de uma peça chegando
    (aba Peças, alimentada pelo robô do porto_tec_panasonic — ver
    routes/pedidos.py:agendar_cliente()), "nossa" é o resto (cliente
    cadastrado na mão, chamado técnico do campo, reagendamento). Mesma
    marca de sempre — EXISTS em pecas_chegada — só que aqui vale pra
    QUALQUER status, não só aguardando_agendamento (ver ?fonte= acima,
    que é a versão específica da fila de Agendar Clientes).
    """
    status = (request.args.get("status") or "").strip()
    cliente_id = request.args.get("cliente_id")
    busca = (request.args.get("busca") or "").strip().lower()
    dias = request.args.get("dias")
    fonte = (request.args.get("fonte") or "").strip().lower()
    origem = (request.args.get("origem") or "").strip().lower()
    pai_id = request.args.get("pai_id")

    condicoes, params = [], []
    if pai_id:
        condicoes.append("os.os_pai_id = ?")
        params.append(pai_id)
    else:
        condicoes.append("os.os_pai_id IS NULL")
    if status:
        condicoes.append("os.status = ?")
        params.append(status)
    if cliente_id:
        condicoes.append("os.cliente_id = ?")
        params.append(cliente_id)
    if dias and str(dias).isdigit() and 1 <= int(dias) <= 3650:
        corte = (datetime.now(timezone.utc) - timedelta(days=int(dias))).strftime("%Y-%m-%d %H:%M:%S")
        condicoes.append("os.criado_em >= ?")
        params.append(corte)
    if fonte == "peca":
        condicoes.append("EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = os.id)")
        condicoes.append("os.oculta_fila_em IS NULL")
    elif fonte == "reagendamento":
        condicoes.append("NOT EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = os.id)")
        condicoes.append("os.oculta_fila_em IS NULL")
    # ?origem=panasonic|nossa|balcao — três lados, pedido de 2026-09-01:
    # "balcao" (aba "Produtos da loja") é o cliente que vem direto na loja,
    # sem passar por Roteiros/técnico nenhum — nunca teve um `servicos`
    # (visita) ligado a essa OS. "nossa" continua sendo o resto de sempre
    # (cliente cadastrado na mão, chamado técnico do campo, reagendamento),
    # agora só que excluindo quem é balcão, pra não aparecer nos dois lados.
    _SEM_ROTEIRO = "NOT EXISTS (SELECT 1 FROM servicos s WHERE s.ordem_servico_id = os.id)"
    if origem == "panasonic":
        condicoes.append("EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = os.id)")
    elif origem == "balcao":
        condicoes.append("NOT EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = os.id)")
        condicoes.append(_SEM_ROTEIRO)
    elif origem == "nossa":
        condicoes.append("NOT EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = os.id)")
        condicoes.append(f"NOT ({_SEM_ROTEIRO})")
    where = f"WHERE {' AND '.join(condicoes)}" if condicoes else ""

    with db_conn() as conn:
        # visita_tecnico/visita_data: pedido de 2026-08-29 ("agendar dia pra
        # atendimento ou aparecer se já tiver") — antes só dava pra ver isso
        # abrindo o detalhe de cada OS uma por uma; agora a própria lista já
        # mostra. Subquery correlacionada (não JOIN) de propósito: uma OS pode
        # ter mais de uma visita ao longo do tempo, e um JOIN duplicaria a
        # linha da OS na lista se houvesse mais de uma pendente.
        ordens = fetch_all(conn, f"""
            SELECT os.*, c.nome AS cliente_nome, c.telefone AS cliente_telefone,
                   (SELECT t.nome
                      FROM servicos s
                      JOIN fichas f ON f.id = s.ficha_id
                      JOIN tecnicos t ON t.id = f.tecnico_id
                     WHERE s.ordem_servico_id = os.id AND s.status = 'pendente'
                     ORDER BY f.data_referencia LIMIT 1) AS visita_tecnico,
                   (SELECT f.data_referencia
                      FROM servicos s
                      JOIN fichas f ON f.id = s.ficha_id
                     WHERE s.ordem_servico_id = os.id AND s.status = 'pendente'
                     ORDER BY f.data_referencia LIMIT 1) AS visita_data
              FROM ordens_servico os
              JOIN clientes c ON c.id = os.cliente_id
              {where}
             ORDER BY os.id DESC
        """, tuple(params))

        # Contagem por status respeita a origem escolhida (senão os números
        # dos cartões não bateriam com a lista de baixo), mas não os outros
        # filtros (status/dias/busca) — mesmo comportamento de sempre.
        _sem_roteiro_contagem = "NOT EXISTS (SELECT 1 FROM servicos s WHERE s.ordem_servico_id = ordens_servico.id)"
        origem_sql = ""
        if origem == "panasonic":
            origem_sql = "WHERE EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = ordens_servico.id)"
        elif origem == "balcao":
            origem_sql = ("WHERE NOT EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = ordens_servico.id) "
                          f"AND {_sem_roteiro_contagem}")
        elif origem == "nossa":
            origem_sql = ("WHERE NOT EXISTS (SELECT 1 FROM pecas_chegada pc WHERE pc.ordem_servico_id = ordens_servico.id) "
                          f"AND NOT ({_sem_roteiro_contagem})")
        contagem = {s: 0 for s in STATUS_OS}
        todas_status = fetch_all(conn, f"SELECT status FROM ordens_servico {origem_sql}")
        for l in todas_status:
            if l["status"] in contagem:
                contagem[l["status"]] += 1

    if busca:
        numero = busca.lstrip("#").lstrip("0") or "0"
        ordens = [
            o for o in ordens
            if busca in (o.get("cliente_nome") or "").lower()
            or (numero.isdigit() and str(o["id"]) == numero)
        ]

    return jsonify({"ordens": ordens, "contagem": contagem, "total": len(ordens)})


@ordens_servico_bp.route("/ordens-servico/exportar", methods=["GET"])
def exportar():
    """Mesmos filtros de listar() (status/dias/busca), em .xlsx — pra
    contabilidade ou reunião, sem precisar copiar linha por linha da tela."""
    status = (request.args.get("status") or "").strip()
    busca = (request.args.get("busca") or "").strip().lower()
    dias = request.args.get("dias")

    condicoes, params = [], []
    if status:
        condicoes.append("os.status = ?")
        params.append(status)
    if dias and str(dias).isdigit() and 1 <= int(dias) <= 3650:
        corte = (datetime.now(timezone.utc) - timedelta(days=int(dias))).strftime("%Y-%m-%d %H:%M:%S")
        condicoes.append("os.criado_em >= ?")
        params.append(corte)
    where = f"WHERE {' AND '.join(condicoes)}" if condicoes else ""

    with db_conn() as conn:
        ordens = fetch_all(conn, f"""
            SELECT os.*, c.nome AS cliente_nome, c.telefone AS cliente_telefone,
                   c.cpf_cnpj AS cliente_cpf_cnpj,
                   t.nome AS tecnico_nome, f.data_referencia, f.dia_semana
              FROM ordens_servico os
              JOIN clientes c ON c.id = os.cliente_id
              LEFT JOIN servicos sv ON sv.id = (
                  SELECT s2.id FROM servicos s2
                   WHERE s2.ordem_servico_id = os.id ORDER BY s2.id DESC LIMIT 1
              )
              LEFT JOIN fichas f ON f.id = sv.ficha_id
              LEFT JOIN tecnicos t ON t.id = f.tecnico_id
              {where}
             ORDER BY os.id DESC
        """, tuple(params))

    if busca:
        numero = busca.lstrip("#").lstrip("0") or "0"
        ordens = [
            o for o in ordens
            if busca in (o.get("cliente_nome") or "").lower()
            or (numero.isdigit() and str(o["id"]) == numero)
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordens de Serviço"

    cabecalho = ["Nº OS", "Status", "Cliente", "CPF/CNPJ", "Telefone",
                "Aparelho", "Marca", "Modelo", "Defeito declarado",
                "Taxa de avaliação", "Atendente", "Técnico", "Dia agendado",
                "Aberta em", "Finalizada em"]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="1A6FD4", end_color="1A6FD4", fill_type="solid")

    for o in ordens:
        dia_agendado = o.get("data_referencia") or o.get("dia_semana") or ""
        ws.append([
            f"OS #{o['id']:06d}", (o["status"] or "").replace("_", " "),
            o.get("cliente_nome") or "", o.get("cliente_cpf_cnpj") or "",
            o.get("cliente_telefone") or "", o.get("tipo_aparelho") or "",
            o.get("marca") or "", o.get("modelo") or "",
            o.get("defeito_declarado") or "", round(o.get("taxa_avaliacao") or 0, 2),
            o.get("atendente") or "", o.get("tecnico_nome") or "", dia_agendado,
            (o.get("criado_em") or "")[:16], (o.get("finalizada_em") or "")[:16],
        ])

    for i, largura in enumerate([11, 20, 22, 16, 15, 14, 12, 14, 30, 12, 14, 14, 13, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"ordens-servico-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buffer, as_attachment=True, download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _pecas_da_os(conn, os_id) -> list:
    """Peças cuja saída de estoque foi vinculada a esta OS — reaproveita
    origem/referencia de estoque_movimentos em vez de criar tabela nova,
    mesmo padrão já usado pra baixa por atendimento."""
    return fetch_all(conn, """
        SELECT m.id, m.quantidade, m.custo_unit, m.criado_em, m.autor,
               e.codigo, e.descricao
          FROM estoque_movimentos m
          JOIN estoque_itens e ON e.id = m.item_id
         WHERE m.origem = 'ordem_servico' AND m.referencia = ? AND m.tipo = 'saida'
         ORDER BY m.id DESC
    """, (str(os_id),))


def _itens_do_orcamento(conn, os_id) -> list:
    return fetch_all(conn, """
        SELECT * FROM ordem_servico_itens WHERE ordem_servico_id = ? ORDER BY id
    """, (os_id,))


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/itens", methods=["GET"])
def listar_itens_orcamento(os_id):
    with db_conn() as conn:
        return jsonify({"itens": _itens_do_orcamento(conn, os_id)})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/itens", methods=["POST"])
def adicionar_item_orcamento(os_id):
    """Linha de serviço+valor no orçamento. ?salvar_catalogo=true (no corpo)
    também grava/atualiza em catalogo_servicos, pro nome aparecer como
    sugestão da próxima vez — sem isso, cada orçamento redigitaria os mesmos
    serviços do zero."""
    d = request.get_json(silent=True) or {}
    nome = (d.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "Informe o nome do serviço"}), 400
    valor = _num(d.get("valor"))

    with db_conn(commit=True) as conn:
        existe = fetch_one(conn, "SELECT id, modelo_os FROM ordens_servico WHERE id = ?", (os_id,))
        if not existe:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404

        item_id = insert_returning_id(conn, """
            INSERT INTO ordem_servico_itens (ordem_servico_id, nome, valor, criado_em)
            VALUES (?, ?, ?, ?)
        """, (os_id, nome, valor, _agora()))

        if d.get("salvar_catalogo"):
            catalogado = fetch_one(conn, "SELECT id FROM catalogo_servicos WHERE LOWER(nome) = LOWER(?)",
                                   (nome,))
            if catalogado:
                execute(conn, "UPDATE catalogo_servicos SET valor = ? WHERE id = ?",
                       (valor, catalogado["id"]))
            else:
                execute(conn, """
                    INSERT INTO catalogo_servicos (nome, valor, criado_em, criado_por)
                    VALUES (?, ?, ?, ?)
                """, (nome, valor, _agora(), _quem()))

        itens = _itens_do_orcamento(conn, os_id)

    return jsonify({"mensagem": "Item adicionado ao orçamento", "id": item_id,
                    "itens": itens}), 201


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/itens/<int:item_id>", methods=["DELETE"])
def remover_item_orcamento(os_id, item_id):
    with db_conn(commit=True) as conn:
        execute(conn, "DELETE FROM ordem_servico_itens WHERE id = ? AND ordem_servico_id = ?",
               (item_id, os_id))
        itens = _itens_do_orcamento(conn, os_id)
    return jsonify({"mensagem": "Item removido", "itens": itens})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/pecas", methods=["GET"])
def listar_pecas(os_id):
    with db_conn() as conn:
        return jsonify({"pecas": _pecas_da_os(conn, os_id)})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/pecas", methods=["POST"])
def adicionar_peca(os_id):
    """Dá saída de uma peça do estoque vinculada a esta OS. Pode ser chamado
    quantas vezes for preciso — um conserto raramente usa uma peça só."""
    d = request.get_json(silent=True) or {}
    codigo = (d.get("codigo") or "").strip()
    if not codigo:
        return jsonify({"erro": "Informe o código da peça"}), 400

    with db_conn(commit=True) as conn:
        existe = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not existe:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404

        try:
            resultado = dar_saida(conn, codigo, d.get("quantidade") or 1,
                                  origem="ordem_servico", referencia=str(os_id),
                                  obs=d.get("obs"))
        except (ValueError, TypeError) as exc:
            return jsonify({"erro": str(exc)}), 400

        # Peça reservada (ver aprovar_orcamento/criar_reserva_os) que agora
        # foi de fato usada — libera a reserva sozinha, sem exigir que
        # alguém lembre de ir lá cancelar manualmente.
        _liberar_reservas_ate(conn, resultado["item_id"], os_id, float(d.get("quantidade") or 1))

    return jsonify({"mensagem": "Peça baixada do estoque e vinculada à OS", **resultado}), 201


# ─── Fotos extras (mais de uma, além da `foto` principal) ────────────────
# Pedido de 2026-08-31. Mesmo padrão já usado pra foto do atendimento em
# campo (servico_foto) — ver services/fotos_extra.py.
@ordens_servico_bp.route("/ordens-servico/<int:os_id>/fotos", methods=["GET"])
def listar_fotos_os(os_id):
    with db_conn() as conn:
        return jsonify({"fotos": listar_fotos_extra(conn, "os", os_id)})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/fotos", methods=["POST"])
def adicionar_foto_os(os_id):
    d = request.get_json(silent=True) or {}
    with db_conn(commit=True) as conn:
        existe = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not existe:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        erro, linha = adicionar_foto_extra(conn, "os", os_id, d.get("foto"))
        if erro:
            return jsonify({"erro": erro}), 400
    return jsonify({"foto": linha}), 201


@ordens_servico_bp.route("/ordens-servico/fotos/<int:foto_id>", methods=["DELETE"])
def apagar_foto_os(foto_id):
    with db_conn(commit=True) as conn:
        if not remover_foto_extra(conn, "os", foto_id):
            return jsonify({"erro": "Foto não encontrada"}), 404
    return jsonify({"mensagem": "Foto removida"})


# ─── Reserva de peça ao aprovar orçamento ─────────────────────────────────
# Pedido de 2026-08-31. Reserva ≠ "Peças usadas" (que já dá baixa de
# verdade): reservar tranca a peça pra ESTA OS sem tirá-la do saldo
# contábil — ninguém vende/usa em outro atendimento enquanto o cliente
# ainda não foi visitado. Vira baixa de verdade (e a reserva se libera
# sozinha) quando a peça é de fato registrada em "Peças usadas" — ver
# adicionar_peca() logo abaixo, que já existia antes desta feature.
def _reservado_do_item(conn, item_id) -> float:
    linha = fetch_one(conn, """
        SELECT COALESCE(SUM(quantidade), 0) AS total FROM estoque_reservas
         WHERE item_id = ? AND liberado_em IS NULL
    """, (item_id,))
    return float(linha["total"] or 0)


def _liberar_reservas_ate(conn, item_id, os_id, quantidade_usada):
    """Libera reservas (mais antiga primeiro) até cobrir o que acabou de ser
    baixado de verdade — libera por INTEIRO, não fraciona reserva."""
    reservas = fetch_all(conn, """
        SELECT id, quantidade FROM estoque_reservas
         WHERE item_id = ? AND ordem_servico_id = ? AND liberado_em IS NULL
         ORDER BY id
    """, (item_id, os_id))
    restante = quantidade_usada
    agora = _agora()
    for r in reservas:
        if restante <= 0:
            break
        execute(conn, "UPDATE estoque_reservas SET liberado_em = ? WHERE id = ?", (agora, r["id"]))
        restante -= float(r["quantidade"])


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/aprovar-orcamento", methods=["PUT"])
def aprovar_orcamento(os_id):
    with db_conn(commit=True) as conn:
        ordem = fetch_one(conn, "SELECT id, modelo_os, orcamento_aprovado_em FROM ordens_servico WHERE id = ?",
                          (os_id,))
        if not ordem:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        if ordem.get("modelo_os") != "orcamento":
            return jsonify({"erro": "Só orçamento pode ser aprovado."}), 400
        if ordem.get("orcamento_aprovado_em"):
            return jsonify({"erro": "Este orçamento já está aprovado."}), 400
        agora = _agora()
        execute(conn, "UPDATE ordens_servico SET orcamento_aprovado_em = ? WHERE id = ?", (agora, os_id))
    return jsonify({"mensagem": "Orçamento aprovado", "orcamento_aprovado_em": agora})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/reservas", methods=["GET"])
def listar_reservas_os(os_id):
    with db_conn() as conn:
        linhas = fetch_all(conn, """
            SELECT r.id, r.quantidade, r.criado_em, r.liberado_em,
                   e.id AS item_id, e.codigo, e.descricao, e.saldo
              FROM estoque_reservas r JOIN estoque_itens e ON e.id = r.item_id
             WHERE r.ordem_servico_id = ?
             ORDER BY r.liberado_em IS NOT NULL, r.id DESC
        """, (os_id,))
    return jsonify({"reservas": linhas})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/reservas", methods=["POST"])
def criar_reserva_os(os_id):
    """Reserva uma peça do estoque pra esta OS — só depois do orçamento
    aprovado (ver aprovar_orcamento acima). Não mexe no saldo contábil, só
    no que está "livre pra prometer" a outro atendimento (ver GET /estoque)."""
    d = request.get_json(silent=True) or {}
    codigo = _norm_codigo(d.get("codigo") or "")
    if not codigo:
        return jsonify({"erro": "Informe o código da peça"}), 400
    try:
        quantidade = float(d.get("quantidade") or 1)
    except (TypeError, ValueError):
        return jsonify({"erro": "Quantidade inválida"}), 400
    if quantidade <= 0:
        return jsonify({"erro": "Quantidade tem que ser maior que zero"}), 400

    with db_conn(commit=True) as conn:
        ordem = fetch_one(conn, "SELECT id, orcamento_aprovado_em FROM ordens_servico WHERE id = ?", (os_id,))
        if not ordem:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        if not ordem.get("orcamento_aprovado_em"):
            return jsonify({"erro": "Aprove o orçamento antes de reservar peça."}), 400

        item = fetch_one(conn, "SELECT * FROM estoque_itens WHERE codigo = ?", (codigo,))
        if not item:
            return jsonify({"erro": f"Peça {codigo} não existe no estoque"}), 404

        disponivel = float(item["saldo"] or 0) - _reservado_do_item(conn, item["id"])
        if quantidade > disponivel:
            return jsonify({
                "erro": f"Só há {disponivel:g} disponível dessa peça (o resto já está "
                        f"reservado por outra OS ou é o próprio saldo)."
            }), 400

        agora = _agora()
        nova_id = insert_returning_id(conn, """
            INSERT INTO estoque_reservas (item_id, ordem_servico_id, quantidade, criado_em)
            VALUES (?, ?, ?, ?)
        """, (item["id"], os_id, quantidade, agora))

    return jsonify({"mensagem": f"{quantidade:g}x {codigo} reservado pra esta OS", "id": nova_id}), 201


@ordens_servico_bp.route("/ordens-servico/reservas/<int:reserva_id>", methods=["DELETE"])
def liberar_reserva_os(reserva_id):
    with db_conn(commit=True) as conn:
        afetadas = execute(conn, """
            UPDATE estoque_reservas SET liberado_em = ?
             WHERE id = ? AND liberado_em IS NULL
        """, (_agora(), reserva_id))
    if not afetadas:
        return jsonify({"erro": "Reserva não encontrada (ou já liberada)"}), 404
    return jsonify({"mensagem": "Reserva liberada"})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>", methods=["GET"])
def obter(os_id):
    with db_conn() as conn:
        os_row = fetch_one(conn, """
            SELECT os.*, c.nome AS cliente_nome, c.telefone AS cliente_telefone,
                   c.email AS cliente_email, c.cpf_cnpj AS cliente_cpf_cnpj,
                   c.endereco AS cliente_endereco, c.numero AS cliente_numero,
                   c.complemento AS cliente_complemento, c.bairro AS cliente_bairro,
                   c.cidade AS cliente_cidade, c.estado AS cliente_estado,
                   c.cep AS cliente_cep
              FROM ordens_servico os
              JOIN clientes c ON c.id = os.cliente_id
             WHERE os.id = ?
        """, (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        try:
            os_row["imprimir_ocultar"] = json.loads(os_row.get("imprimir_ocultar") or "[]")
        except (TypeError, ValueError):
            os_row["imprimir_ocultar"] = []

        visitas = fetch_all(conn, """
            SELECT s.id, s.status, s.ordem, f.id AS ficha_id, f.dia_semana,
                   f.data_referencia, t.id AS tecnico_id, t.nome AS tecnico_nome,
                   t.cor AS tecnico_cor, d.desfecho, d.observacao AS desfecho_obs
              FROM servicos s
              JOIN fichas f ON f.id = s.ficha_id
              LEFT JOIN tecnicos t ON t.id = f.tecnico_id
              LEFT JOIN servico_desfecho d ON d.servico_id = s.id
             WHERE s.ordem_servico_id = ?
             ORDER BY s.id DESC
        """, (os_id,))

        pecas = _pecas_da_os(conn, os_id)
        # Itens (Serviço/Peças/Mão de obra) não são mais exclusivos do
        # Orçamento — qualquer OS pode ter, ver criar().
        itens = _itens_do_orcamento(conn, os_id)

        # Chamado Técnico/Orçamento pendurados nesta OS (ver os_pai_id) — só
        # existe quando ESTA é uma OS de primeiro nível; uma filha não tem filha.
        filhas = [] if os_row.get("os_pai_id") else fetch_all(conn, """
            SELECT id, modelo_os, status, defeito_declarado, criado_em
              FROM ordens_servico WHERE os_pai_id = ? ORDER BY id DESC
        """, (os_id,))

    termos = TERMOS_POR_TIPO.get(os_row.get("tipo_os"), TERMOS_PADRAO)
    return jsonify({"ordem": os_row, "visitas": visitas, "pecas": pecas,
                    "itens": itens, "filhas": filhas, "termos": termos})


@ordens_servico_bp.route("/ordens-servico", methods=["POST"])
def criar():
    """Body: {cliente_id} OU {cliente_novo: {...}} — abrir OS com cliente que
    já existe, ou cadastrar e abrir na mesma tacada (é o caminho mais comum:
    cliente novo ligando pela primeira vez).

    ?modelo_os decide qual conjunto de campos vale (ver MODELOS_OS): 'os'
    (padrão) EXIGE tipo_os pro termo jurídico; 'chamado_tecnico' e
    'orcamento' aceitam solucao/foto/tecnico_atendeu_id e, no caso do
    orçamento, uma lista inicial de itens — tipo_os aí é OPCIONAL (dá pra
    amarrar um termo de garantia a um orçamento, por exemplo, mas não é
    obrigatório like na OS padrão).
    """
    d = request.get_json(silent=True) or {}
    campos = _campos_os(d)

    if not campos["tipo_aparelho"] and not campos["defeito_declarado"]:
        return jsonify({"erro": "Informe ao menos o aparelho ou o defeito declarado"}), 400

    erro_modelo, modelo_os = _validar_modelo_os(d.get("modelo_os"))
    if erro_modelo:
        return jsonify({"erro": erro_modelo}), 400

    # Setor obrigatório em "Ordens de Serviço" e "Chamado Técnico" — pedido de
    # 2026-08-31, mesma régua que já existe no atendimento de Roteiros
    # (routes/servicos.py): sem classificar por fabricante o relatório por
    # setor fica capenga pra quem entra pela OS. Orçamento fica de fora — não
    # foi pedido, e nem sempre tem um fabricante definido ainda.
    setor_id = None
    if modelo_os in ("os", "chamado_tecnico"):
        erro_setor, setor_id = _validar_setor(d.get("setor_id"))
        if erro_setor:
            return jsonify({"erro": erro_setor}), 400

    tipo_os = None
    if modelo_os == "os":
        erro_tipo_os, tipo_os = _validar_tipo_os(d.get("tipo_os"))
        if erro_tipo_os:
            return jsonify({"erro": erro_tipo_os}), 400
    else:
        tipo_os_bruto = (d.get("tipo_os") or "").strip()
        if tipo_os_bruto:
            if tipo_os_bruto not in TIPOS_OS:
                return jsonify({"erro": "Tipo de OS inválido."}), 400
            tipo_os = tipo_os_bruto

    solucao = (d.get("solucao") or "").strip()
    forma_pagamento = (d.get("forma_pagamento") or "").strip()
    # Só o modelo Chamado Técnico usa — vale 0 pros outros dois, que cobram
    # por Itens/Valores em vez de uma taxa fixa de vistoria.
    taxa_vistoria = _num(d.get("taxa_vistoria"))
    ocultar_fila = bool(d.get("oculta_fila"))
    imprimir_ocultar = _validar_imprimir_ocultar(d.get("imprimir_ocultar"))
    # Garantia vale pra "saída da oficina" (data calculada pelo tipo) e pro
    # modelo Orçamento (o cliente pode querer já deixar combinado um prazo,
    # mesmo sem ainda ter um tipo_os fechado) — pedido de 2026-08-29.
    _usa_garantia = tipo_os == "saida_oficina" or modelo_os == "orcamento"
    garantia_inicio = _validar_data_iso(d.get("garantia_inicio")) if _usa_garantia else None
    garantia_meses = _validar_garantia_meses(d.get("garantia_meses")) if _usa_garantia else None
    # Foto/técnico deixaram de ser exclusivos do Chamado Técnico — pedido de
    # 2026-08-27, mesma OS pode precisar registrar isso em qualquer modelo.
    foto = _foto_valida(d.get("foto"))
    try:
        tecnico_atendeu_id = int(d["tecnico_atendeu_id"]) if d.get("tecnico_atendeu_id") else None
    except (TypeError, ValueError):
        tecnico_atendeu_id = None

    with db_conn(commit=True) as conn:
        cliente_id = d.get("cliente_id")
        if not cliente_id:
            cliente_novo = d.get("cliente_novo") or {}
            try:
                cliente_id = criar_cliente(conn, cliente_novo)
            except ValueError as exc:
                return jsonify({"erro": str(exc)}), 400
        else:
            existe = fetch_one(conn, "SELECT id FROM clientes WHERE id = ?", (cliente_id,))
            if not existe:
                return jsonify({"erro": "Cliente não encontrado"}), 404

        # Pendura num caso já aberto do MESMO cliente em vez de virar número
        # solto — pedido de 2026-08-27, pra Chamado Técnico/Orçamento de
        # quem já tem OS não empilhar linha nova a cada atendimento. Só
        # aceita pai de primeiro nível (sem os_pai_id ele mesmo) e do mesmo
        # cliente — não faz sentido pendurar OS de gente diferente.
        os_pai_id = None
        if d.get("os_pai_id"):
            try:
                candidato_pai_id = int(d["os_pai_id"])
            except (TypeError, ValueError):
                return jsonify({"erro": "OS pai inválida"}), 400
            pai = fetch_one(conn, "SELECT id, cliente_id, os_pai_id FROM ordens_servico WHERE id = ?",
                            (candidato_pai_id,))
            if not pai or pai["cliente_id"] != int(cliente_id) or pai.get("os_pai_id"):
                return jsonify({"erro": "OS pai inválida"}), 400
            os_pai_id = pai["id"]

        agora = _agora()
        oculta_fila_em = agora if ocultar_fila else None
        # Toda OS ganha um link público próprio (token_cliente) — não é só a
        # criada pelo técnico em campo que pode ser mandada pro cliente
        # (ver app.py:/os/cliente/<token>); qualquer uma pode, a qualquer
        # momento.
        token_cliente = secrets.token_urlsafe(24)
        os_id = insert_returning_id(conn, """
            INSERT INTO ordens_servico
                (cliente_id, atendente, tipo_aparelho, marca, modelo,
                 numero_serie, voltagem, acessorios, defeito_declarado, taxa_avaliacao,
                 status, observacao, criado_em, criado_por, tipo_os,
                 modelo_os, solucao, foto, tecnico_atendeu_id, os_pai_id, forma_pagamento,
                 token_cliente, taxa_vistoria, oculta_fila_em, imprimir_ocultar, garantia_inicio,
                 garantia_meses, setor_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (cliente_id, _quem(), campos["tipo_aparelho"], campos["marca"],
              campos["modelo"], campos["numero_serie"], campos["voltagem"], campos["acessorios"],
              campos["defeito_declarado"], _num(d.get("taxa_avaliacao")),
              "aguardando_agendamento", campos["observacao"], agora, _quem(),
              tipo_os, modelo_os, solucao, foto, tecnico_atendeu_id, os_pai_id, forma_pagamento,
              token_cliente, taxa_vistoria, oculta_fila_em, imprimir_ocultar, garantia_inicio,
              garantia_meses, setor_id))

        # Itens (Serviço/Peças/Mão de obra) não são mais exclusivos do
        # Orçamento — pedido de 2026-08-27, baseado no modelo impresso que
        # tem essa seção pra qualquer OS. Aceito em qualquer modelo.
        for item in (d.get("itens") or []):
            nome_item = (item.get("nome") or "").strip()
            if not nome_item:
                continue
            execute(conn, """
                INSERT INTO ordem_servico_itens (ordem_servico_id, nome, valor, criado_em)
                VALUES (?, ?, ?, ?)
            """, (os_id, nome_item, _num(item.get("valor")), agora))

    return jsonify({"mensagem": "Ordem de serviço aberta", "id": os_id,
                    "cliente_id": cliente_id}), 201


@ordens_servico_bp.route("/ordens-servico/<int:os_id>", methods=["PUT"])
def editar(os_id):
    d = request.get_json(silent=True) or {}

    with db_conn(commit=True) as conn:
        existe = fetch_one(conn, "SELECT id, modelo_os, tipo_os FROM ordens_servico WHERE id = ?", (os_id,))
        if not existe:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        # Efetivo = o modelo/tipo que a OS vai TER depois deste PUT (se vier
        # no corpo) — decide se tipo_os é obrigatório e se garantia_inicio
        # ainda faz sentido guardar.
        modelo_efetivo = (d.get("modelo_os") or existe.get("modelo_os") or "os").strip()
        tipo_os_efetivo = (d.get("tipo_os") if "tipo_os" in d else existe.get("tipo_os")) or ""

        campos, valores = [], []
        for chave in ("tipo_aparelho", "marca", "modelo", "numero_serie", "voltagem",
                     "acessorios", "defeito_declarado", "observacao", "solucao",
                     "forma_pagamento"):
            if chave in d:
                campos.append(f"{chave} = ?")
                valores.append((d.get(chave) or "").strip())
        if "taxa_avaliacao" in d:
            campos.append("taxa_avaliacao = ?")
            valores.append(_num(d.get("taxa_avaliacao")))
        if "taxa_vistoria" in d:
            campos.append("taxa_vistoria = ?")
            valores.append(_num(d.get("taxa_vistoria")))
        if "imprimir_ocultar" in d:
            campos.append("imprimir_ocultar = ?")
            valores.append(_validar_imprimir_ocultar(d.get("imprimir_ocultar")))
        if "status" in d:
            status = (d.get("status") or "").strip()
            if status not in STATUS_OS:
                return jsonify({"erro": f"Status inválido. Use um de: {', '.join(STATUS_OS)}"}), 400
            campos.append("status = ?")
            valores.append(status)
            if status == "finalizada":
                campos.append("finalizada_em = ?")
                valores.append(_agora())
        if "tipo_os" in d:
            if modelo_efetivo == "os":
                erro_tipo_os, tipo_os = _validar_tipo_os(d.get("tipo_os"))
                if erro_tipo_os:
                    return jsonify({"erro": erro_tipo_os}), 400
            else:
                # Opcional fora do modelo "Ordens de Serviço": dá pra deixar
                # em branco (limpa o termo) sem travar o salvamento.
                tipo_os_bruto = (d.get("tipo_os") or "").strip()
                if tipo_os_bruto and tipo_os_bruto not in TIPOS_OS:
                    return jsonify({"erro": "Tipo de OS inválido."}), 400
                tipo_os = tipo_os_bruto or None
            campos.append("tipo_os = ?")
            valores.append(tipo_os)
        if "modelo_os" in d:
            # Retroativo, de propósito: OS aberta antes desse recurso existir
            # (ou aberta como "Ordens de Serviço" por engano) pode virar
            # Chamado Técnico ou Orçamento depois — não é só quem entra
            # daqui pra frente que ganha os três modelos.
            erro_modelo, modelo_os = _validar_modelo_os(d.get("modelo_os"))
            if erro_modelo:
                return jsonify({"erro": erro_modelo}), 400
            campos.append("modelo_os = ?")
            valores.append(modelo_os)
        if "foto" in d:
            campos.append("foto = ?")
            valores.append(_foto_valida(d.get("foto")))
        if "tecnico_atendeu_id" in d:
            try:
                tecnico_atendeu_id = int(d["tecnico_atendeu_id"]) if d.get("tecnico_atendeu_id") else None
            except (TypeError, ValueError):
                tecnico_atendeu_id = None
            campos.append("tecnico_atendeu_id = ?")
            valores.append(tecnico_atendeu_id)
        if "setor_id" in d:
            # Classificação por fabricante — pedido de 2026-08-31, mesma
            # régua de servicos.py. Só entra quando mandado explicitamente:
            # diferente da garantia, setor não é amarrado a modelo/tipo, então
            # trocar de modelo não apaga o que já foi classificado.
            erro_setor, setor_id = _validar_setor(d.get("setor_id"))
            if erro_setor:
                return jsonify({"erro": erro_setor}), 400
            campos.append("setor_id = ?")
            valores.append(setor_id)
        # "saida_oficina" e o modelo Orçamento usam garantia — pedido de
        # 2026-08-29 pra poder já deixar combinado um prazo no orçamento.
        _usa_garantia_efetivo = tipo_os_efetivo == "saida_oficina" or modelo_efetivo == "orcamento"
        if "garantia_inicio" in d or (("tipo_os" in d or "modelo_os" in d) and not _usa_garantia_efetivo):
            # Fora desses dois casos, guardar essa data não tem pra que servir
            # na impressão (ver _GARANTIA_MESES em app.py). O segundo caso
            # (trocou de tipo_os/modelo_os sem mandar garantia_inicio junto)
            # limpa sozinho — senão a data velha ficava enterrada no banco e
            # reaparecia se o tipo/modelo voltasse a valer garantia.
            campos.append("garantia_inicio = ?")
            valores.append(_validar_data_iso(d.get("garantia_inicio"))
                           if _usa_garantia_efetivo else None)
        if "garantia_meses" in d or (("tipo_os" in d or "modelo_os" in d) and not _usa_garantia_efetivo):
            campos.append("garantia_meses = ?")
            valores.append(_validar_garantia_meses(d.get("garantia_meses"))
                           if _usa_garantia_efetivo else None)

        if not campos:
            return jsonify({"mensagem": "Nada para mudar"})

        campos.append("atualizado_em = ?")
        valores.append(_agora())
        valores.append(os_id)
        execute(conn, f"UPDATE ordens_servico SET {', '.join(campos)} WHERE id = ?", valores)

    return jsonify({"mensagem": "Ordem de serviço atualizada"})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>", methods=["DELETE"])
def apagar(os_id):
    """Apaga a OS de verdade — pedido de 2026-08-27 (só dava pra criar,
    nunca pra remover). Itens (ordem_servico_itens) somem junto por
    ON DELETE CASCADE; visitas (servicos.ordem_servico_id) e filhas
    (os_pai_id) só perdem a referência (ON DELETE SET NULL) — apagar uma OS
    não pode apagar atendimento que aconteceu de verdade nem outra OS."""
    with db_conn(commit=True) as conn:
        existe = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not existe:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        execute(conn, "DELETE FROM ordens_servico WHERE id = ?", (os_id,))
    return jsonify({"mensagem": "Ordem de serviço apagada"})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/pedir-peca", methods=["POST"])
def pedir_peca(os_id):
    """Botão "Pedir peça" no detalhe da OS — pedido de 2026-08-28. Diferente
    do "Precisa de peça" que o técnico marca em campo (servico_desfecho,
    preso a um atendimento): aqui é o escritório pedindo peça pra uma OS
    que pode nem ter visita marcada ainda. Cai junto na aba Atendimentos
    (ver listar_desfechos em routes/relatorios.py, que faz UNION dos dois)."""
    d = request.get_json(silent=True) or {}
    peca = (d.get("peca") or "").strip()
    descricao = (d.get("descricao") or "").strip()
    if not peca and not descricao:
        return jsonify({"erro": "Informe ao menos a peça ou a descrição"}), 400
    foto = _foto_valida(d.get("foto"))

    with db_conn(commit=True) as conn:
        os_row = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        pedido_id = insert_returning_id(conn, """
            INSERT INTO pedido_peca_os
                (ordem_servico_id, peca, descricao, foto, criado_em, criado_por)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (os_id, peca, descricao, foto, _agora(), _quem()))

    return jsonify({"mensagem": "Peça pedida — já aparece em Atendimentos", "id": pedido_id}), 201


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/link-cliente", methods=["GET"])
def link_cliente(os_id):
    """Devolve o token do link público desta OS, gerando um na hora se ela
    for de antes desse recurso existir (2026-08-28) e ainda não tiver."""
    with db_conn(commit=True) as conn:
        os_row = fetch_one(conn, "SELECT id, token_cliente FROM ordens_servico WHERE id = ?", (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        token = os_row.get("token_cliente")
        if not token:
            token = secrets.token_urlsafe(24)
            execute(conn, "UPDATE ordens_servico SET token_cliente = ? WHERE id = ?", (token, os_id))
    return jsonify({"token": token})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/agendar", methods=["POST"])
def agendar(os_id):
    """Cria (ou reaproveita) uma visita na agenda do técnico escolhido.

    Body: {tecnico_id, ficha_id} pra um dia que já existe, ou
    {tecnico_id, nova_data} pra abrir um dia novo — mesmo mecanismo do
    reagendamento em tecnico_api.py, só que aqui quem escolhe o técnico é
    quem está abrindo a OS, não o próprio técnico.
    """
    d = request.get_json(silent=True) or {}
    try:
        tecnico_id = int(d.get("tecnico_id"))
    except (TypeError, ValueError):
        return jsonify({"erro": "Escolha o técnico"}), 400

    with db_conn(commit=True) as conn:
        os_row = fetch_one(conn, """
            SELECT os.*, c.nome AS cliente_nome, c.cep AS cliente_cep,
                   c.endereco AS cliente_endereco, c.numero AS cliente_numero,
                   c.bairro AS cliente_bairro, c.cidade AS cliente_cidade
              FROM ordens_servico os JOIN clientes c ON c.id = os.cliente_id
             WHERE os.id = ?
        """, (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404

        tecnico = fetch_one(conn, "SELECT id FROM tecnicos WHERE id = ?", (tecnico_id,))
        if not tecnico:
            return jsonify({"erro": "Técnico não encontrado"}), 404

        ficha_id = d.get("ficha_id")
        if ficha_id:
            ficha = fetch_one(conn, "SELECT * FROM fichas WHERE id = ? AND tecnico_id = ?",
                              (ficha_id, tecnico_id))
            if not ficha:
                return jsonify({"erro": "Esse dia não existe ou não é desse técnico"}), 404
        else:
            nova_data = (d.get("nova_data") or "").strip()
            if not nova_data:
                return jsonify({"erro": "Escolha um dia existente ou uma data nova"}), 400
            from routes.fichas import nome_dia_semana
            try:
                dia = nome_dia_semana(nova_data)
            except ValueError:
                return jsonify({"erro": "Data inválida"}), 400

            # Travado contra duas OS sendo agendadas pro mesmo técnico/dia
            # quase ao mesmo tempo criarem ficha duplicada — obter_ou_criar_ficha
            # em routes/fichas.py é o ponto único que resolve isso pros três
            # lugares que precisam de "reaproveita se existe, senão cria".
            ficha_id, _ = obter_ou_criar_ficha(conn, tecnico_id, dia, nova_data)
            ficha = fetch_one(conn, "SELECT * FROM fichas WHERE id = ?", (ficha_id,))

        # Endereço do cliente vira o ponto na rota. Sem CEP não geocodifica,
        # mas a visita ainda entra — corrige depois é melhor que não poder
        # agendar por falta de um dado que a etiqueta às vezes não tem.
        cep = "".join(c for c in (os_row.get("cliente_cep") or "") if c.isdigit())
        geo = geocode_cep(cep, numero=os_row.get("cliente_numero") or "") if cep else None
        endereco_completo = (geo.endereco if geo else " ".join(x for x in [
            os_row.get("cliente_endereco"), os_row.get("cliente_numero"),
            os_row.get("cliente_bairro"), os_row.get("cliente_cidade"),
        ] if x)) or None

        ultima = fetch_one(conn, "SELECT MAX(ordem) AS m FROM servicos WHERE ficha_id = ?",
                           (ficha_id,))
        servico_id = insert_returning_id(conn, """
            INSERT INTO servicos (ficha_id, cep, numero, endereco_completo, lat, lng,
                                  cliente, descricao, ordem, status, tipo_aparelho,
                                  modelo, ordem_servico_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (ficha_id, cep, os_row.get("cliente_numero") or "", endereco_completo,
              geo.lat if geo else None, geo.lng if geo else None,
              os_row.get("cliente_nome"), os_row.get("defeito_declarado"),
              ((ultima or {}).get("m") or 0) + 1, "pendente",
              os_row.get("tipo_aparelho"), os_row.get("modelo"), os_id))

        execute(conn, "UPDATE ordens_servico SET status = ?, atualizado_em = ? WHERE id = ?",
               ("agendada", _agora(), os_id))

        recalcular_rota(conn, ficha_id, ficha)

    return jsonify({"mensagem": "Visita agendada", "ficha_id": ficha_id,
                    "servico_id": servico_id}), 201


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/desagendar/<int:servico_id>", methods=["DELETE"])
def desagendar(os_id, servico_id):
    """Desfaz um agendamento feito com técnico/dia errado.

    Só mexe em visita ainda PENDENTE — uma já concluída tem desfecho e
    histórico reais, desagendar isso apagaria trabalho que aconteceu de
    verdade. A visita errada é removida (nunca chegou a acontecer, não há o
    que preservar); se não sobrar nenhuma outra visita pendente ou
    concluída, a OS volta pra 'aguardando_agendamento'.
    """
    with db_conn(commit=True) as conn:
        servico = fetch_one(conn, """
            SELECT * FROM servicos WHERE id = ? AND ordem_servico_id = ?
        """, (servico_id, os_id))
        if not servico:
            return jsonify({"erro": "Visita não encontrada nesta OS"}), 404
        if servico["status"] != "pendente":
            return jsonify({"erro": "Só dá pra desagendar uma visita que ainda não aconteceu"}), 400

        ficha_id = servico["ficha_id"]
        execute(conn, "DELETE FROM servicos WHERE id = ?", (servico_id,))

        restantes = fetch_one(conn, """
            SELECT COUNT(*) AS total FROM servicos WHERE ordem_servico_id = ?
        """, (os_id,))["total"]
        if not restantes:
            execute(conn, "UPDATE ordens_servico SET status = ?, atualizado_em = ? WHERE id = ?",
                   ("aguardando_agendamento", _agora(), os_id))

        ficha = fetch_one(conn, "SELECT * FROM fichas WHERE id = ?", (ficha_id,))
        if ficha:
            recalcular_rota(conn, ficha_id, ficha)

    return jsonify({"mensagem": "Visita desagendada"})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/ocultar-fila", methods=["POST"])
def ocultar_fila(os_id):
    """Some com o cartão em Agendar Clientes sem mexer no status da OS nem
    apagar nada — pedido de 2026-08-27, pra tirar da vista um caso já
    resolvido por fora (cliente ligou direto, por exemplo) sem forçar um
    status que mentiria sobre o que aconteceu de verdade. Reversível via
    /reexibir-fila: a OS continua inteira, só marcada pra não aparecer ali.
    """
    with db_conn(commit=True) as conn:
        os_row = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        execute(conn, "UPDATE ordens_servico SET oculta_fila_em = ? WHERE id = ?",
               (_agora(), os_id))
    return jsonify({"mensagem": "Removida da fila de Agendar Clientes"})


@ordens_servico_bp.route("/ordens-servico/<int:os_id>/reexibir-fila", methods=["POST"])
def reexibir_fila(os_id):
    """Desfaz o /ocultar-fila — a OS volta a aparecer em Agendar Clientes se
    ainda estiver em 'aguardando_agendamento'."""
    with db_conn(commit=True) as conn:
        os_row = fetch_one(conn, "SELECT id FROM ordens_servico WHERE id = ?", (os_id,))
        if not os_row:
            return jsonify({"erro": "Ordem de serviço não encontrada"}), 404
        execute(conn, "UPDATE ordens_servico SET oculta_fila_em = NULL WHERE id = ?", (os_id,))
    return jsonify({"mensagem": "De volta pra fila de Agendar Clientes"})
