"""Substituição de peças da Panasonic — pedido de 2026-08-29.

A Panasonic manda de vez em quando uma planilha (SAP) dizendo que o código
X foi substituído pelos códigos Y, Z... Isso é DIFERENTE da planilha de
Pedidos (routes/pedidos.py): aquela é sobre COMPRA (o que já foi pedido,
pra quem é); esta é um catálogo de referência (o que virou o quê), que só
muda quando a fabricante reedita a lista inteira.

Por isso o fluxo é upload manual (a planilha mora no computador de quem
sobe, não em nenhuma pasta que o servidor consiga ler sozinho) que
SUBSTITUI a tabela inteira, e não faz merge — tentar casar linha por linha
entre duas edições da mesma planilha, sem uma chave estável confirmada, é
mais chance de erro do que começar do zero a cada import.
"""
import io
import logging
import os
from datetime import datetime, timezone

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook

from database import IS_PG, db_conn, execute, fetch_all, fetch_one, sql

log = logging.getLogger("portotec.substituicoes")

substituicoes_bp = Blueprint("substituicoes", __name__)

# Colunas da planilha da Panasonic (1-indexed): E = código original,
# I..M = até 5 substitutos em cadeia, F/G = validade. As quatro primeiras
# (Aplicação/Tipo de determinação/Organização/Canal) não interessam aqui —
# são campos internos do SAP da fabricante, sempre iguais na prática.
COL_CODIGO = 5
COL_INICIO_VALIDADE = 6
COL_FIM_VALIDADE = 7
COL_SUBSTITUTOS = (9, 10, 11, 12, 13)


def _normalizar_codigo(texto):
    return (texto or "").strip().upper()


def _texto_data(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    return str(valor).strip() or None


def _agora():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _token_valido():
    esperado = os.environ.get("PANASONIC_SYNC_TOKEN")
    return bool(esperado) and request.headers.get("X-Sync-Token") == esperado


def _pendente_bool(linha_cache):
    """Nenhuma linha em precos_panasonic ainda == igual a pendente (acabou de
    entrar na fila agora mesmo, _marcar_pendente_se_novo roda DEPOIS dessa
    leitura). SQLite guarda pendente como 0/1, Postgres como bool nativo —
    normaliza os dois pra um bool limpo no JSON."""
    if linha_cache is None:
        return True
    return bool(linha_cache.get("pendente"))


def _precos_do_cache(conn, codigos):
    """{codigo: {preco, atualizado_em, pendente}} pros códigos pedidos — só os
    que já têm alguma linha em precos_panasonic (pendente ou não). `pendente`
    é o que diferencia "robô ainda não verificou" de "robô já verificou e a
    Panasonic não tem preço pra mostrar" — sem isso os dois casos pareciam a
    mesma coisa ("consultando...") pro técnico, e um preço ausente permanente
    parecia estar travado pra sempre (confusão relatada em 2026-09-01)."""
    if not codigos:
        return {}
    marcador = ",".join("?" for _ in codigos)
    linhas = fetch_all(conn, f"""
        SELECT codigo, preco, atualizado_em, pendente FROM precos_panasonic
         WHERE codigo IN ({marcador})
    """, tuple(codigos))
    return {l["codigo"]: l for l in linhas}


def _marcar_pendente_se_novo(conn, codigos):
    """Um código pesquisado que nunca foi consultado antes vira pendente —
    é a fila que o robô local (Portotec/Softwear para Pedidos) processa
    primeiro, priorizando o que alguém pediu de verdade em vez de varrer a
    planilha inteira do zero."""
    agora = _agora()
    for codigo in codigos:
        existe = fetch_one(conn, "SELECT codigo FROM precos_panasonic WHERE codigo = ?", (codigo,))
        if not existe:
            execute(conn, """
                INSERT INTO precos_panasonic (codigo, pendente, solicitado_em)
                VALUES (?, ?, ?)
            """, (codigo, True if IS_PG else 1, agora))


@substituicoes_bp.route("/pecas-substituicao/status", methods=["GET"])
def status():
    with db_conn() as conn:
        total = fetch_one(conn, "SELECT COUNT(*) AS n FROM pecas_substituicao")["n"]
    return jsonify({"total": total})


@substituicoes_bp.route("/pecas-substituicao", methods=["GET"])
def buscar():
    """Devolve os substitutos do código pesquisado. Casamento exato primeiro
    (é o caso normal — código de peça não é texto livre); se não achar nada,
    tenta por "contém" — cobre quem copiou o código com espaço ou um
    caractere a mais/a menos colado por engano."""
    codigo = _normalizar_codigo(request.args.get("codigo"))
    if not codigo:
        return jsonify({"erro": "Informe o código da peça"}), 400

    hoje = datetime.now().strftime("%Y-%m-%d")
    with db_conn() as conn:
        linhas = fetch_all(conn, sql("""
            SELECT * FROM pecas_substituicao
             WHERE codigo = ?
               AND (fim_validade IS NULL OR fim_validade >= ?)
        """), (codigo, hoje))
        casamento = "exato"
        # Sem filtro de validade: um formato de data que a planilha trouxe
        # diferente do esperado não pode fazer um código que EXISTE parecer
        # "não encontrado" — melhor mostrar vencido do que esconder de vez.
        if not linhas:
            linhas = fetch_all(conn, "SELECT * FROM pecas_substituicao WHERE codigo = ?", (codigo,))
        if not linhas:
            linhas = fetch_all(conn, sql("""
                SELECT * FROM pecas_substituicao
                 WHERE codigo LIKE ?
                 LIMIT 20
            """), (f"%{codigo}%",))
            casamento = "parcial"

    resultados = []
    # O código digitado SEMPRE entra na fila de preço, tenha ou não substituto
    # cadastrado na planilha (pedido de 2026-09-01) — antes só entrava se
    # desse "match" na tabela de substituição, então a maioria das peças
    # (que não têm substituto documentado) nunca recebia preço nenhum.
    todos_codigos = {codigo}
    for l in linhas:
        substitutos = [l[f"substituto_{i}"] for i in range(1, 6) if l.get(f"substituto_{i}")]
        if not substitutos:
            continue
        resultados.append({
            "codigo": l["codigo"],
            "substitutos": substitutos,
            "inicio_validade": l.get("inicio_validade"),
            "fim_validade": l.get("fim_validade"),
        })
        todos_codigos.add(l["codigo"])
        todos_codigos.update(substitutos)

    # Preço Panasonic B2B: cache que o robô local mantém (ver
    # _marcar_pendente_se_novo/POST /precos-panasonic mais abaixo — o site
    # não consegue abrir aquele portal sozinho, quem consulta de verdade é
    # o computador do Kalebe). Pesquisar aqui é o que ENTRA um código na
    # fila desse robô pela primeira vez.
    with db_conn(commit=True) as conn:
        precos = _precos_do_cache(conn, list(todos_codigos))
        _marcar_pendente_se_novo(conn, list(todos_codigos))

    for r in resultados:
        cache_principal = precos.get(r["codigo"])
        r["preco_panasonic"] = cache_principal.get("preco") if cache_principal else None
        r["preco_atualizado_em"] = cache_principal.get("atualizado_em") if cache_principal else None
        r["preco_pendente"] = _pendente_bool(cache_principal)
        r["substitutos_precos"] = [
            {"codigo": s, "preco": (precos.get(s) or {}).get("preco"),
             "pendente": _pendente_bool(precos.get(s))}
            for s in r["substitutos"]
        ]

    cache_busca = precos.get(codigo)
    return jsonify({
        "codigo_buscado": codigo,
        "casamento": casamento,
        "resultados": resultados,
        # Preço do PRÓPRIO código digitado, independente de ter substituto
        # cadastrado — é o que faltava pra maioria das peças ter preço.
        "preco_panasonic": cache_busca.get("preco") if cache_busca else None,
        "preco_atualizado_em": cache_busca.get("atualizado_em") if cache_busca else None,
        "preco_pendente": _pendente_bool(cache_busca),
    })


@substituicoes_bp.route("/precos-panasonic/pendentes", methods=["GET"])
def precos_pendentes():
    """Fila pro robô local processar — protegida por token compartilhado
    (não é login de usuário, é máquina falando com máquina)."""
    if not _token_valido():
        return jsonify({"erro": "Token inválido"}), 401
    limite = min(int(request.args.get("limite", 30) or 30), 200)
    with db_conn() as conn:
        pendentes = fetch_all(conn, sql("""
            SELECT codigo FROM precos_panasonic
             WHERE pendente = ?
             ORDER BY solicitado_em
             LIMIT ?
        """), (True if IS_PG else 1, limite))
    return jsonify({"codigos": [p["codigo"] for p in pendentes]})


@substituicoes_bp.route("/precos-panasonic", methods=["POST"])
def gravar_precos():
    """O robô local manda o resultado de volta — {itens: [{codigo, preco}]}.
    `preco` None/vazio marca como consultado mas sem preço disponível (não
    fica pendente pra sempre tentando de novo a cada rodada)."""
    if not _token_valido():
        return jsonify({"erro": "Token inválido"}), 401
    d = request.get_json(silent=True) or {}
    itens = d.get("itens") or []
    if not itens:
        return jsonify({"erro": "Mande 'itens': [{codigo, preco}]"}), 400

    # UPDATE-e-só-se-preciso-INSERT em vez de ON CONFLICT: mesmo motivo do
    # bump_revisao em database.py — ON CONFLICT tem diferença de dialeto
    # entre SQLite e Postgres, e aqui não há necessidade de arriscar isso.
    agora = _agora()
    pendente_falso = False if IS_PG else 0
    with db_conn(commit=True) as conn:
        for item in itens:
            codigo = _normalizar_codigo(item.get("codigo"))
            if not codigo:
                continue
            preco = (item.get("preco") or "").strip() or None
            afetadas = execute(conn, """
                UPDATE precos_panasonic SET preco = ?, pendente = ?, atualizado_em = ?
                 WHERE codigo = ?
            """, (preco, pendente_falso, agora, codigo))
            if not afetadas:
                execute(conn, """
                    INSERT INTO precos_panasonic (codigo, preco, pendente, atualizado_em)
                    VALUES (?, ?, ?, ?)
                """, (codigo, preco, pendente_falso, agora))

    return jsonify({"mensagem": f"{len(itens)} preço(s) atualizado(s)"})


@substituicoes_bp.route("/pecas-substituicao/importar", methods=["POST"])
def importar():
    """Sobe a planilha da Panasonic e SUBSTITUI a tabela inteira por ela."""
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Anexe o arquivo .xlsx"}), 400
    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"erro": "Só aceito .xlsx (o mesmo formato que a Panasonic manda)"}), 400

    try:
        conteudo = arquivo.read()
        # Trava de tamanho: sem isso, um arquivo errado gigante trava o
        # processo tentando carregar tudo em memória de uma vez.
        if len(conteudo) > 25 * 1024 * 1024:
            return jsonify({"erro": "Arquivo maior que 25MB — confira se é a planilha certa"}), 400
        wb = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        aba = wb[wb.sheetnames[0]]
    except Exception as exc:
        log.exception("Falha ao abrir a planilha de substituição")
        return jsonify({"erro": f"Não consegui abrir o arquivo: {exc}"}), 400

    linhas = []
    for i, linha in enumerate(aba.iter_rows(min_row=2, values_only=True)):
        if not linha or len(linha) < COL_CODIGO:
            continue
        codigo = _normalizar_codigo(linha[COL_CODIGO - 1])
        if not codigo:
            continue
        substitutos = []
        for col in COL_SUBSTITUTOS:
            valor = linha[col - 1] if len(linha) >= col else None
            substitutos.append(_normalizar_codigo(valor) or None)
        inicio = _texto_data(linha[COL_INICIO_VALIDADE - 1] if len(linha) >= COL_INICIO_VALIDADE else None)
        fim = _texto_data(linha[COL_FIM_VALIDADE - 1] if len(linha) >= COL_FIM_VALIDADE else None)
        linhas.append((codigo, *substitutos, inicio, fim))

    if not linhas:
        return jsonify({"erro": "Nenhuma linha reconhecida nesse arquivo — confira se é a planilha certa"}), 400

    try:
        with db_conn(commit=True) as conn:
            execute(conn, "DELETE FROM pecas_substituicao")
            cur = conn.cursor()
            try:
                if IS_PG:
                    # executemany do psycopg2 manda uma INSERT por linha — 18 mil
                    # viagens de ida e volta ao Postgres do Railway estouram fácil
                    # o --timeout 60 do gunicorn (Procfile) e o import morre no
                    # meio, sem avisar direito: a busca depois só dizia "não
                    # achou nada". execute_values manda tudo numa única instrução.
                    from psycopg2.extras import execute_values
                    execute_values(cur, """
                        INSERT INTO pecas_substituicao
                            (codigo, substituto_1, substituto_2, substituto_3, substituto_4,
                             substituto_5, inicio_validade, fim_validade)
                        VALUES %s
                    """, linhas, page_size=2000)
                else:
                    cur.executemany(sql("""
                        INSERT INTO pecas_substituicao
                            (codigo, substituto_1, substituto_2, substituto_3, substituto_4,
                             substituto_5, inicio_validade, fim_validade)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """), linhas)
            finally:
                cur.close()
    except Exception as exc:
        # Antes disso, uma falha aqui virava 500 genérico do Flask — sem
        # mensagem nenhuma pra quem subiu o arquivo, só "não achou nada" na
        # busca depois, sem ligar os dois fatos. Log pra investigar server-side
        # (não expõe a mensagem crua pra evitar vazar detalhe de conexão do
        # banco, mas confirma que FALHOU, que é o que faltava ficar claro).
        log.exception("Falha ao gravar a planilha de substituição no banco")
        return jsonify({"erro": f"Consegui ler o arquivo ({len(linhas)} linhas), "
                                f"mas falhou ao gravar no banco: {type(exc).__name__}. "
                                f"Tenta de novo — se continuar, me avisa."}), 502

    # Confere que gravou de verdade — depois do bug do executemany travando
    # sem avisar, "a chamada não deu erro" deixou de ser garantia suficiente
    # de que os dados chegaram no banco.
    with db_conn() as conn:
        total_no_banco = fetch_one(conn, "SELECT COUNT(*) AS n FROM pecas_substituicao")["n"]
    if total_no_banco != len(linhas):
        log.error("Import de substituição: esperava %d linhas, banco ficou com %d",
                  len(linhas), total_no_banco)
        return jsonify({"erro": f"Gravação incompleta: esperava {len(linhas)} linhas, "
                                f"o banco ficou com {total_no_banco}. Tenta importar de novo."}), 502

    return jsonify({"mensagem": "Planilha importada", "total": len(linhas)}), 201
