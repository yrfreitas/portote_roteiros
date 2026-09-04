import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from database import (db_conn, execute, fetch_all, fetch_one,
                      insert_returning_id, sql)

relatorios_bp = Blueprint("relatorios", __name__)

log = logging.getLogger("portotec.relatorios")

FOTO_MAXIMA = 900 * 1024
PREFIXOS_FOTO = ("data:image/jpeg;base64,", "data:image/png;base64,",
                 "data:image/webp;base64,")


def _foto_valida(foto):
    if not isinstance(foto, str) or not foto.startswith(PREFIXOS_FOTO):
        return None
    if len(foto) > FOTO_MAXIMA:
        return None
    return foto


def _parse_dt(valor):
    if not valor:
        return None
    texto = str(valor).strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


@relatorios_bp.route("/historico/exportar", methods=["GET"])
def exportar_historico():
    # Acompanha o filtro de período da tela. Sem isso, quem estivesse olhando
    # os últimos 7 dias exportaria a planilha inteira e levaria para uma
    # reunião um número que não confere com o que viu.
    # Mesmo formato de corte do listar_fichas: espaço como separador, porque a
    # coluna é TEXT gravada com CURRENT_TIMESTAMP e a comparação é textual.
    dias = request.args.get("dias")
    filtro, params = "", ()
    if dias and str(dias).isdigit() and 1 <= int(dias) <= 3650:
        corte = datetime.now(timezone.utc) - timedelta(days=int(dias))
        filtro = "AND f.concluida_em >= ?"
        params = (corte.strftime("%Y-%m-%d %H:%M:%S"),)

    with db_conn() as conn:
        fichas = fetch_all(conn, f"""
            SELECT f.dia_semana, f.data_referencia, f.concluida_em,
                   f.distancia_total, t.nome AS tecnico_nome,
                   COUNT(s.id) AS total_servicos
            FROM fichas f
            LEFT JOIN servicos s ON s.ficha_id = f.id
            LEFT JOIN tecnicos t ON t.id = f.tecnico_id
            WHERE f.status = 'concluida' {filtro}
            GROUP BY f.id, t.nome
            ORDER BY f.concluida_em DESC
        """, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"

    cabecalho = ["Técnico", "Dia da semana", "Data de referência",
                 "Concluída em", "Pontos atendidos", "Km rodados"]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="1A6FD4", end_color="1A6FD4", fill_type="solid")

    for f in fichas:
        concluida_em = _parse_dt(f.get("concluida_em"))
        ws.append([
            f.get("tecnico_nome") or "—",
            f.get("dia_semana") or "",
            f.get("data_referencia") or "",
            concluida_em.strftime("%d/%m/%Y %H:%M") if concluida_em else "",
            f.get("total_servicos") or 0,
            round(f.get("distancia_total") or 0, 1),
        ])

    for i, largura in enumerate([22, 16, 16, 18, 16, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"historico-rotas-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    from flask import session

    from database import registrar_exportacao
    with db_conn(commit=True) as conn:
        registrar_exportacao(conn, session.get("usuario_nome") or "",
                             "/historico/exportar", f"dias={dias or 'todos'}")

    return send_file(
        buffer, as_attachment=True, download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@relatorios_bp.route("/metricas/tendencia", methods=["GET"])
def metricas_tendencia():
    """Agrega rotas concluídas por técnico e por semana (últimas 8 semanas).
    Feito em Python em vez de SQL de data porque SQLite e Postgres não
    compartilham as mesmas funções de data — mais portável assim."""
    with db_conn() as conn:
        fichas = fetch_all(conn, """
            SELECT f.tecnico_id, t.nome AS tecnico_nome, t.cor AS tecnico_cor,
                   f.concluida_em, f.distancia_total,
                   COUNT(s.id) AS total_servicos
            FROM fichas f
            LEFT JOIN servicos s ON s.ficha_id = f.id
            LEFT JOIN tecnicos t ON t.id = f.tecnico_id
            WHERE f.status = 'concluida' AND f.concluida_em IS NOT NULL
            GROUP BY f.id, t.nome, t.cor
        """)

    hoje = datetime.now()
    inicio_janela = hoje - timedelta(weeks=8)

    semanas = defaultdict(lambda: defaultdict(lambda: {"rotas": 0, "pontos": 0, "km": 0.0}))
    tecnicos_vistos = {}

    for f in fichas:
        dt = _parse_dt(f.get("concluida_em"))
        if not dt or dt < inicio_janela:
            continue

        semana_iso = f"{dt.isocalendar()[0]}-S{dt.isocalendar()[1]:02d}"
        tid = f["tecnico_id"]
        tecnicos_vistos[tid] = {"nome": f.get("tecnico_nome") or "—", "cor": f.get("tecnico_cor") or "#4f8dfb"}

        bucket = semanas[semana_iso][tid]
        bucket["rotas"] += 1
        bucket["pontos"] += f.get("total_servicos") or 0
        bucket["km"] += f.get("distancia_total") or 0

    semanas_ordenadas = sorted(semanas.keys())

    resultado = []
    for semana in semanas_ordenadas:
        por_tecnico = []
        for tid, dados in semanas[semana].items():
            info = tecnicos_vistos.get(tid, {"nome": "—", "cor": "#4f8dfb"})
            por_tecnico.append({
                "tecnico_id": tid,
                "tecnico_nome": info["nome"],
                "tecnico_cor": info["cor"],
                "rotas": dados["rotas"],
                "pontos": dados["pontos"],
                "km": round(dados["km"], 1),
            })
        por_tecnico.sort(key=lambda x: x["km"], reverse=True)
        resultado.append({"semana": semana, "tecnicos": por_tecnico})

    return jsonify({"semanas": resultado})


@relatorios_bp.route("/relatorios/tecnicos", methods=["GET"])
def comparativo_tecnicos():
    """Comparativo entre técnicos: carga, quilometragem e conclusão.

    Passou a fazer sentido quando a Porto Tec chegou a dois técnicos em campo
    (2026-08-17). Serve para duas perguntas concretas: quem está sobrecarregado
    e quanto cada rota custa em estrada.

    Conta TODAS as fichas, concluídas ou não — a de tendência já olha só as
    fechadas, e aqui o que interessa é a carga real de cada um agora.
    ?dias=N limita pela data de referência (padrão: tudo em aberto + 60 dias).
    """
    try:
        dias = max(1, min(365, int(request.args.get("dias", 60))))
    except (TypeError, ValueError):
        dias = 60

    corte = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")

    with db_conn() as conn:
        linhas = fetch_all(conn, """
            SELECT t.id, t.nome, t.cor, t.foto,
                   f.id AS ficha_id, f.status, f.distancia_total,
                   f.data_referencia,
                   COUNT(s.id) AS pontos,
                   SUM(CASE WHEN s.status = 'concluido' THEN 1 ELSE 0 END) AS concluidos
              FROM tecnicos t
              LEFT JOIN fichas f ON f.tecnico_id = t.id
              LEFT JOIN servicos s ON s.ficha_id = f.id
             GROUP BY t.id, t.nome, t.cor, t.foto, f.id, f.status,
                      f.distancia_total, f.data_referencia
             ORDER BY t.nome
        """)

    por_tecnico = {}
    for l in linhas:
        # Fichas antigas ficam de fora, mas o TÉCNICO continua aparecendo —
        # some da lista quem não existe, não quem teve um mês parado.
        data = l.get("data_referencia") or ""
        dentro = (not data) or (data >= corte)

        t = por_tecnico.setdefault(l["id"], {
            "id": l["id"], "nome": l["nome"], "cor": l.get("cor") or "#4f8dfb",
            "foto": l.get("foto"),
            "rotas": 0, "rotas_concluidas": 0, "pontos": 0,
            "concluidos": 0, "km": 0.0,
        })

        if not l.get("ficha_id") or not dentro:
            continue

        t["rotas"] += 1
        if l.get("status") == "concluida":
            t["rotas_concluidas"] += 1
        t["pontos"] += l.get("pontos") or 0
        t["concluidos"] += l.get("concluidos") or 0
        t["km"] += float(l.get("distancia_total") or 0)

    resultado = []
    for t in por_tecnico.values():
        t["km"] = round(t["km"], 1)
        # Km POR PONTO é o número que compara de verdade: quilometragem alta
        # com muitos pontos é rota cheia; alta com poucos é rota espalhada,
        # que é o que dói no combustível.
        t["km_por_ponto"] = round(t["km"] / t["pontos"], 1) if t["pontos"] else 0
        t["taxa_conclusao"] = (round(100 * t["concluidos"] / t["pontos"])
                               if t["pontos"] else 0)
        t["pendentes"] = t["pontos"] - t["concluidos"]
        resultado.append(t)

    resultado.sort(key=lambda x: -x["pontos"])

    total_pontos = sum(t["pontos"] for t in resultado) or 1
    for t in resultado:
        t["fatia"] = round(100 * t["pontos"] / total_pontos)

    return jsonify({"dias": dias, "tecnicos": resultado,
                    "total_pontos": total_pontos})


# ─── Painel de atendimentos registrados pelo técnico ────────────────────
#
# O técnico passou a dizer o que aconteceu em cada visita (resolvido, precisa
# de peça, volta depois, não atendido) e a mandar a foto da etiqueta. Sem uma
# tela que junte isso, o dado fica espalhado ponto a ponto dentro de cada
# rota — existe, mas ninguém enxerga.
#
# É aqui que "precisa de peça" vira lista de trabalho: são os atendimentos que
# dependem de alguém comprar alguma coisa para poderem terminar.

# ─── "Alguém já tem essa peça no carro?" ────────────────────────────────
#
# É a pergunta que precisa ser feita ANTES de comprar. Sem ela o escritório
# pedia peça que já estava rodando na van de um técnico, e o cliente esperava
# a compra chegar enquanto a peça passava na porta dele.
#
# O casamento é pelo CÓDIGO, exato. Descrição varia entre o que o técnico
# digita e o que está no estoque; código não.
RE_CODIGO = re.compile(r"([A-Z][A-Z0-9]{2,}(?:-[A-Z0-9]+)?)")


def _codigos(texto: str) -> set:
    return {c for c in RE_CODIGO.findall((texto or "").upper())
            if any(ch.isdigit() for ch in c)}


def _marcar_disponivel_no_carro(conn, atendimentos):
    """Anota em cada atendimento que precisa de peça quem já a carrega."""
    pendentes = [a for a in atendimentos
                 if a.get("desfecho") == "precisa_peca" and a.get("peca")]
    if not pendentes:
        return

    estoque = fetch_all(conn, sql("""
        SELECT c.codigo, c.quantidade, c.tecnico_id, t.nome, t.cor
          FROM peca_carro c JOIN tecnicos t ON t.id = c.tecnico_id
         WHERE c.quantidade > 0
    """))
    if not estoque:
        return

    por_codigo = {}
    for e in estoque:
        por_codigo.setdefault(e["codigo"], []).append(e)

    for a in pendentes:
        achados = []
        for codigo in _codigos(a["peca"]):
            for e in por_codigo.get(codigo, []):
                achados.append({"codigo": codigo, "tecnico": e["nome"],
                                "tecnico_id": e["tecnico_id"],
                                "cor": e["cor"], "quantidade": e["quantidade"]})
        if achados:
            a["no_carro"] = achados


DESFECHOS_ORDEM = ["precisa_peca", "cotacao_peca", "volto_depois", "nao_atendido",
                   "fazer_os", "resolvido"]


@relatorios_bp.route("/desfechos", methods=["GET"])
def listar_desfechos():
    """Atendimentos com desfecho registrado, com contagem por tipo.

    ?dias=N limita o período (padrão 30). ?tipo=X filtra um desfecho.
    """
    try:
        dias = max(1, min(365, int(request.args.get("dias", 30))))
    except (TypeError, ValueError):
        dias = 30

    tipo = (request.args.get("tipo") or "").strip()
    limite = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d 00:00:00")

    with db_conn() as conn:
        linhas = fetch_all(conn, sql("""
            SELECT d.servico_id, d.desfecho, d.motivo, d.peca, d.observacao,
                   d.pedido_em, d.pedido_por,
                   d.registrado_em, d.registrado_por,
                   s.cliente, s.endereco_completo, s.tipo_aparelho, s.modelo,
                   s.numero_os, s.ficha_id, s.ordem_servico_id,
                   f.dia_semana, f.data_referencia,
                   t.nome AS tecnico, t.cor AS tecnico_cor,
                   (SELECT COUNT(*) FROM servico_foto sf
                     WHERE sf.servico_id = d.servico_id) AS fotos
              FROM servico_desfecho d
              JOIN servicos s ON s.id = d.servico_id
              LEFT JOIN fichas f ON f.id = s.ficha_id
              LEFT JOIN tecnicos t ON t.id = f.tecnico_id
             WHERE d.registrado_em >= ?
             ORDER BY d.registrado_em DESC
        """), (limite,))
        for l in linhas:
            l["origem"] = "tecnico"
            l["pedido_os_id"] = None
            l["chave"] = f"t{l['servico_id']}"

        # "Pedir peça" batido direto na OS (sem visita de técnico envolvida) —
        # mesma vitrine de Atendimentos, mas sem servico_desfecho por trás (ver
        # pedido_peca_os em database.py: servico_desfecho.servico_id é PRIMARY
        # KEY, não dá pra pendurar um pedido de peça sem ficha ali).
        pecas_os = fetch_all(conn, sql("""
            SELECT p.id, p.peca, p.descricao, p.foto AS peca_foto,
                   p.criado_em, p.criado_por,
                   p.pedido_em, p.pedido_por, p.pedido_foto,
                   os.id AS ordem_servico_id, os.tipo_aparelho, os.modelo,
                   c.nome AS cliente
              FROM pedido_peca_os p
              JOIN ordens_servico os ON os.id = p.ordem_servico_id
              JOIN clientes c ON c.id = os.cliente_id
             WHERE p.criado_em >= ?
             ORDER BY p.criado_em DESC
        """), (limite,))
        for p in pecas_os:
            linhas.append({
                "servico_id": None, "desfecho": "precisa_peca", "motivo": None,
                "peca": p["peca"], "observacao": p["descricao"],
                "pedido_em": p["pedido_em"], "pedido_por": p["pedido_por"],
                "pedido_foto": p["pedido_foto"],
                "registrado_em": p["criado_em"], "registrado_por": p["criado_por"],
                "cliente": p["cliente"], "endereco_completo": None,
                "tipo_aparelho": p["tipo_aparelho"], "modelo": p["modelo"],
                "numero_os": None, "ficha_id": None,
                "ordem_servico_id": p["ordem_servico_id"],
                "dia_semana": None, "data_referencia": None,
                "tecnico": None, "tecnico_cor": None, "fotos": 0,
                "peca_foto": p["peca_foto"],
                "origem": "os", "pedido_os_id": p["id"], "chave": f"o{p['id']}",
            })
        linhas.sort(key=lambda l: l["registrado_em"] or "", reverse=True)

    # Contagem vem do conjunto INTEIRO do período, antes de aplicar o filtro:
    # os números do topo têm de continuar mostrando o total de cada tipo mesmo
    # quando a lista abaixo está filtrada em um deles.
    contagem = {k: 0 for k in DESFECHOS_ORDEM}
    for l in linhas:
        if l["desfecho"] in contagem:
            contagem[l["desfecho"]] += 1

    with db_conn() as conn:
        _marcar_disponivel_no_carro(conn, linhas)

    if tipo in contagem:
        linhas = [l for l in linhas if l["desfecho"] == tipo]

    return jsonify({
        "dias": dias,
        "total": sum(contagem.values()),
        "contagem": contagem,
        "atendimentos": linhas,
    })


@relatorios_bp.route("/desfechos/<int:servico_id>/pedido", methods=["POST"])
def marcar_peca_pedida(servico_id):
    """Dá baixa: a peça daquele atendimento foi pedida.

    Faz duas coisas de uma vez, e é essa junção que fecha o circuito:
    marca o atendimento aqui e ESCREVE NA PLANILHA quem é o cliente e qual a
    peça. Assim o pedido deixa de existir só na cabeça de quem comprou.

    Se a planilha falhar, a baixa AINDA acontece e a resposta avisa. O
    contrário — travar a baixa porque o Google recusou — deixaria o operador
    sem saber se pediu ou não, que é pior que uma linha faltando na planilha.
    """
    from flask import session

    data = request.get_json(silent=True) or {}
    foto = data.get("foto")
    if foto is not None:
        foto = _foto_valida(foto)
        if not foto:
            return jsonify({"erro": "Foto inválida ou grande demais."}), 400

    with db_conn() as conn:
        dado = fetch_one(conn, sql("""
            SELECT d.servico_id, d.desfecho, d.peca, d.observacao, d.pedido_em,
                   s.cliente, s.numero_os, s.tipo_aparelho, s.modelo,
                   t.nome AS tecnico
              FROM servico_desfecho d
              JOIN servicos s ON s.id = d.servico_id
              LEFT JOIN fichas f ON f.id = s.ficha_id
              LEFT JOIN tecnicos t ON t.id = f.tecnico_id
             WHERE d.servico_id = ?
        """), (servico_id,))

    if not dado:
        return jsonify({"erro": "Atendimento não encontrado"}), 404
    if dado["pedido_em"]:
        return jsonify({"erro": "Esta peça já foi marcada como pedida.",
                        "pedido_em": dado["pedido_em"]}), 409

    quem = (session.get("usuario_nome") or "").strip()[:80]
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with db_conn(commit=True) as conn:
        execute(conn, sql("UPDATE servico_desfecho SET pedido_em = ?, "
                          "pedido_por = ?, pedido_foto = ? WHERE servico_id = ?"),
                (agora, quem, foto, servico_id))

    aviso = None
    try:
        from services.planilha import registrar_peca_solicitada
        r = registrar_peca_solicitada({
            "servico_id": servico_id,
            "cliente": dado.get("cliente"),
            "peca": dado.get("peca"),
            "aparelho": " ".join(x for x in [dado.get("tipo_aparelho"),
                                             dado.get("modelo")] if x),
            "numero_os": dado.get("numero_os"),
            "tecnico": dado.get("tecnico"),
            "observacao": dado.get("observacao"),
            "pedido_por": quem,
        })
        if not r.get("configurada"):
            aviso = "Baixa registrada, mas a planilha não está configurada."
    except Exception as exc:
        log.exception("Falha ao gravar peça solicitada na planilha")
        aviso = f"Baixa registrada, mas não consegui escrever na planilha: {exc}"

    return jsonify({"pedido_em": agora, "pedido_por": quem, "tem_foto": bool(foto), "aviso": aviso})


@relatorios_bp.route("/desfechos/<int:servico_id>/pedido", methods=["DELETE"])
def desfazer_peca_pedida(servico_id):
    """Desfaz a baixa — pedido de 2026-09-01: comprovante errado anexado
    sem querer, precisa voltar pra fila de "Precisa de peça" (Atendimentos)
    pra corrigir e pedir de novo. NÃO apaga a linha já escrita na planilha
    do Panasonic (se a baixa tiver chegado até lá) — isso é manual, fica um
    registro a mais lá, inofensivo."""
    with db_conn(commit=True) as conn:
        afetadas = execute(conn, sql(
            "UPDATE servico_desfecho SET pedido_em = NULL, pedido_por = NULL, "
            "pedido_foto = NULL WHERE servico_id = ?"), (servico_id,))
    if not afetadas:
        return jsonify({"erro": "Atendimento não encontrado"}), 404
    return jsonify({"mensagem": "Pedido desfeito — volta a aparecer em Atendimentos"})


@relatorios_bp.route("/pedidos-peca-os/<int:pedido_id>/pedido", methods=["DELETE"])
def desfazer_peca_os_pedida(pedido_id):
    """Mesma coisa, pro pedido de peça batido direto na OS (sem visita de
    técnico) — ver marcar_peca_os_pedida logo abaixo."""
    with db_conn(commit=True) as conn:
        afetadas = execute(conn, sql(
            "UPDATE pedido_peca_os SET pedido_em = NULL, pedido_por = NULL, "
            "pedido_foto = NULL WHERE id = ?"), (pedido_id,))
    if not afetadas:
        return jsonify({"erro": "Pedido não encontrado"}), 404
    return jsonify({"mensagem": "Pedido desfeito"})


@relatorios_bp.route("/desfechos/<int:servico_id>", methods=["DELETE"])
def apagar_desfecho(servico_id):
    """Tira uma linha de vez da lista de Atendimentos — pedido de 2026-09-02,
    entrou coisa errada em "Precisam de peça" sem jeito de remover.

    Apaga só o DESFECHO (o que foi registrado sobre a visita), não a visita
    nem a OS por trás — a linha some daqui, o resto do sistema não muda."""
    with db_conn(commit=True) as conn:
        afetadas = execute(conn, sql(
            "DELETE FROM servico_desfecho WHERE servico_id = ?"), (servico_id,))
    if not afetadas:
        return jsonify({"erro": "Atendimento não encontrado"}), 404
    return jsonify({"mensagem": "Removido de Atendimentos"})


@relatorios_bp.route("/pedidos-peca-os/<int:pedido_id>", methods=["DELETE"])
def apagar_pedido_peca_os(pedido_id):
    """Mesma coisa, pro pedido de peça batido direto na OS (sem visita)."""
    with db_conn(commit=True) as conn:
        afetadas = execute(conn, sql(
            "DELETE FROM pedido_peca_os WHERE id = ?"), (pedido_id,))
    if not afetadas:
        return jsonify({"erro": "Pedido não encontrado"}), 404
    return jsonify({"mensagem": "Removido de Atendimentos"})


@relatorios_bp.route("/pedidos-peca-os/manual", methods=["POST"])
def criar_pedido_peca_manual():
    """Pedido de peça criado NA MÃO, sem vir de um désfecho de técnico
    (pedido de 2026-09-02, revisado no mesmo dia depois do pedido de mover
    isso pra "Pedidos com comprovante"). Body: {peca, descricao, foto}
    + OU {cliente_id} OU {cliente_novo: {nome, telefone}} OU nenhum dos
    dois (reposição de estoque, sem cliente nenhum).

    Não fabrica mais uma OS pra pendurar o pedido: pedido_peca_os.
    ordem_servico_id é NULLABLE desde a migração de 2026-09-02 justamente
    pra isso — pedido sem cliente não tem do que ser "ordem de serviço".
    Cliente_id (também novo na tabela) vai direto, sem OS no meio.

    Foto é OBRIGATÓRIA (diferente do pedido que nasce de désfecho de
    técnico, onde dá pra marcar "precisa de peça" sem foto e completar
    depois): sem OS por trás, um pedido sem foto não tem em lugar nenhum
    pra reaparecer e ser completado — nasceria pendente e órfão pra
    sempre. Já entra marcado como pedido (pedido_em/pedido_por/pedido_foto)."""
    from flask import session

    from routes.clientes import criar_cliente

    dados = request.get_json(silent=True) or {}
    peca = (dados.get("peca") or "").strip()[:200]
    if not peca:
        return jsonify({"erro": "Informe a peça"}), 400
    descricao = (dados.get("descricao") or "").strip()[:600]
    quem = (session.get("usuario_nome") or "").strip()[:80] or "Administrador"
    agora = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    # Foto OBRIGATÓRIA aqui (diferente da cotação de peça): esta tela é
    # "Pedidos com comprovante" — sem OS por trás, um pedido sem foto não
    # tem em lugar nenhum pra reaparecer depois pra alguém completar.
    foto = _foto_valida(dados.get("foto"))
    if not foto:
        return jsonify({"erro": "Anexe a foto do comprovante da compra"}), 400

    with db_conn(commit=True) as conn:
        cliente_id = dados.get("cliente_id")
        cliente_novo = dados.get("cliente_novo")
        if cliente_id:
            existe = fetch_one(conn, sql("SELECT id FROM clientes WHERE id = ?"), (cliente_id,))
            if not existe:
                return jsonify({"erro": "Cliente não encontrado"}), 404
        elif cliente_novo:
            try:
                cliente_id = criar_cliente(conn, cliente_novo)
            except ValueError as exc:
                return jsonify({"erro": str(exc)}), 400
        else:
            cliente_id = None   # reposição de estoque — sem cliente mesmo

        pedido_id = insert_returning_id(conn, sql("""
            INSERT INTO pedido_peca_os
                (cliente_id, peca, descricao, criado_em, criado_por,
                 pedido_em, pedido_por, pedido_foto)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """), (cliente_id, peca, descricao, agora, quem, agora, quem, foto))

    return jsonify({"mensagem": "Pedido criado", "id": pedido_id}), 201


@relatorios_bp.route("/relatorios/sugestao-preco", methods=["GET"])
def sugestao_preco():
    """Preço médio já APROVADO pelo cliente pra esse tipo de aparelho —
    pedido de 2026-09-03: "precificação inteligente do orçamento".

    NÃO é machine learning nem IA — é média/mediana em cima do histórico
    real de `ordem_servico_itens` de OS cujo orçamento já foi aprovado
    (orcamento_aprovado_em preenchido). Sem histórico nenhum daquele
    aparelho, devolve `n: 0` e quem chama decide não mostrar nada — um
    valor sozinho, de UM caso, não é sugestão confiável, é ruído.
    """
    aparelho = (request.args.get("aparelho") or "").strip()
    if not aparelho:
        return jsonify({"erro": "Informe o aparelho"}), 400

    with db_conn() as conn:
        linhas = fetch_all(conn, sql("""
            SELECT i.valor FROM ordem_servico_itens i
              JOIN ordens_servico os ON os.id = i.ordem_servico_id
             WHERE os.orcamento_aprovado_em IS NOT NULL
                   AND LOWER(os.tipo_aparelho) = LOWER(?)
        """), (aparelho,))

    valores = sorted(float(l["valor"]) for l in linhas if l.get("valor"))
    if len(valores) < 3:   # menos que isso não é padrão, é coincidência
        return jsonify({"n": len(valores)})

    media = sum(valores) / len(valores)
    mediana = valores[len(valores) // 2]
    return jsonify({
        "n": len(valores), "media": round(media, 2), "mediana": round(mediana, 2),
        "minimo": round(valores[0], 2), "maximo": round(valores[-1], 2),
    })


@relatorios_bp.route("/backup", methods=["GET"])
def baixar_backup():
    """Backup sob demanda, baixado na hora — pedido de 2026-09-02: "backup
    automático diário do banco, com botão de restaurar um ponto no tempo".

    Escopo reduzido de propósito: o AGENDAMENTO diário e o RESTORE por um
    clique ficam de fora desta rodada — restaurar sozinho numa ferramenta
    de auto-serviço é perigoso demais num sistema de produção (um clique
    errado sobrescreveria dados reais de cliente/OS sem confirmação
    nenhuma no meio). O que entra é a parte segura e imediatamente útil:
    baixar um retrato completo do banco agora, em JSON, pra guardar antes
    de mexer em algo arriscado ou só por precaução. Restaurar, se um dia
    precisar, é o Kalebe rodando um script à parte com o arquivo — não um
    botão que qualquer sessão logada pode apertar sem querer.

    Tabelas de sistema descobertas pelo próprio catálogo do banco
    (pg_tables/sqlite_master), não uma lista fixa no código — uma tabela
    nova de amanhã já entra no backup de amanhã sem precisar editar isto.
    """
    from permissoes import pode

    if not pode("gerenciar_usuarios"):
        return jsonify({"erro": "Sem permissão"}), 403

    import json

    from database import IS_PG

    with db_conn() as conn:
        if IS_PG:
            tabelas = fetch_all(conn, sql(
                "SELECT tablename AS nome FROM pg_tables WHERE schemaname = 'public'"))
        else:
            tabelas = fetch_all(conn, sql(
                "SELECT name AS nome FROM sqlite_master WHERE type = 'table'"))

        dump = {}
        for t in tabelas:
            nome = t["nome"]
            if nome.startswith("sqlite_"):
                continue
            try:
                dump[nome] = fetch_all(conn, f"SELECT * FROM {nome}")
            except Exception as exc:
                log.warning("Backup: falha lendo tabela %s: %s", nome, exc)
                dump[nome] = {"erro_ao_ler": str(exc)}

    conteudo = json.dumps(dump, default=str, ensure_ascii=False, indent=2).encode("utf-8")
    nome_arquivo = f"backup_portotec_{datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
    return send_file(io.BytesIO(conteudo), mimetype="application/json",
                     as_attachment=True, download_name=nome_arquivo)


@relatorios_bp.route("/relatorios/resumo-dia", methods=["GET"])
def resumo_dia():
    """Resumo do dia em frases prontas — pedido de 2026-09-03: "resumo
    automático do dia... visão do dia sem precisar caçar em várias abas".

    NÃO é gerado por IA (a ANTHROPIC_API_KEY do servidor está com a chave
    inválida agora — 401 authentication_error, achado ao investigar este
    mesmo pedido; consertar isso é decisão de fora do código, precisa de
    chave nova). É montado por regra, com número de verdade de cada
    consulta — funciona hoje, sem depender de nada externo, e dá pra trocar
    por um resumo redigido pela IA depois sem mudar o formato que o front
    espera (lista de frases).
    """
    hoje = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    tres_dias_atras = (datetime.now(timezone.utc) - timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")

    with db_conn() as conn:
        fechadas_hoje = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM ordens_servico
             WHERE finalizada_em IS NOT NULL AND finalizada_em >= ?
        """), (f"{hoje} 00:00:00",))

        orcamentos_parados = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM ordens_servico
             WHERE status IN ('aguardando_orcamento', 'aguardando_aprovacao')
                   AND COALESCE(atualizado_em, criado_em) <= ?
        """), (tres_dias_atras,))

        precisam_peca = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM servico_desfecho
             WHERE desfecho = 'precisa_peca' AND pedido_em IS NULL
        """))

        abaixo_minimo = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM estoque_itens WHERE minimo > 0 AND saldo <= minimo
        """))

        cotacoes_pendentes = fetch_one(conn, sql(
            "SELECT COUNT(*) AS n FROM cotacoes WHERE status = 'pendente'"))

    n_fechadas = fechadas_hoje["n"] or 0
    n_parados = orcamentos_parados["n"] or 0
    n_peca = precisam_peca["n"] or 0
    n_minimo = abaixo_minimo["n"] or 0
    n_cotacao = cotacoes_pendentes["n"] or 0

    frases = []
    frases.append(f"{n_fechadas} OS fechada{'s' if n_fechadas != 1 else ''} hoje."
                  if n_fechadas else "Nenhuma OS fechada hoje ainda.")
    if n_parados:
        frases.append(f"{n_parados} orçamento{'s' if n_parados != 1 else ''} parado{'s' if n_parados != 1 else ''} "
                      f"há 3+ dias sem aprovação nem preço.")
    if n_peca:
        frases.append(f"{n_peca} atendimento{'s' if n_peca != 1 else ''} esperando comprar peça.")
    if n_cotacao:
        frases.append(f"{n_cotacao} peça{'s' if n_cotacao != 1 else ''} na fila de cotação sem preço ainda.")
    if n_minimo:
        frases.append(f"{n_minimo} peça{'s' if n_minimo != 1 else ''} no estoque abaixo do mínimo.")
    if len(frases) == 1 and not (n_parados or n_peca or n_cotacao or n_minimo):
        frases.append("Sem pendência parada nas filas de orçamento, peça ou estoque agora.")

    return jsonify({
        "frases": frases,
        # Números crus pra quem quiser ligar isso a um KPI (Central de
        # Comando) sem precisar caçar dentro do texto da frase com regex
        # -- o jeito antigo (ver git blame) quebrava se a frase mudasse.
        "numeros": {
            "fechadas": n_fechadas, "orcamentos_parados": n_parados,
            "precisa_peca": n_peca, "estoque_minimo": n_minimo,
            "cotacoes_pendentes": n_cotacao,
        },
    })


@relatorios_bp.route("/relatorios/comparativo-dia", methods=["GET"])
def comparativo_dia():
    """Hoje vs ontem, pra Central de Comando mostrar seta pra cima/baixo
    nos KPIs em vez de só o número seco -- pedido de 2026-09-04.
    """
    hoje = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ontem = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y-%m-%d")

    with db_conn() as conn:
        fechadas_hoje = fetch_one(conn, sql(
            "SELECT COUNT(*) AS n FROM ordens_servico WHERE finalizada_em >= ?"), (f"{hoje} 00:00:00",))
        fechadas_ontem = fetch_one(conn, sql(
            "SELECT COUNT(*) AS n FROM ordens_servico WHERE finalizada_em >= ? AND finalizada_em < ?"),
            (f"{ontem} 00:00:00", f"{hoje} 00:00:00"))
        atendimentos_hoje = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM servicos sv JOIN fichas f ON f.id = sv.ficha_id
             WHERE f.data_referencia = ?"""), (hoje,))
        atendimentos_ontem = fetch_one(conn, sql("""
            SELECT COUNT(*) AS n FROM servicos sv JOIN fichas f ON f.id = sv.ficha_id
             WHERE f.data_referencia = ?"""), (ontem,))

    return jsonify({
        "fechadas": {"hoje": fechadas_hoje["n"] or 0, "ontem": fechadas_ontem["n"] or 0},
        "atendimentos": {"hoje": atendimentos_hoje["n"] or 0, "ontem": atendimentos_ontem["n"] or 0},
    })


_DESFECHO_ROTULO_FEED = {
    "resolvido": "resolveu o atendimento",
    "orcamento": "levantou orçamento",
    "precisa_peca": "pediu peça",
    "volto_depois": "vai voltar depois",
    "cotacao_peca": "pediu cotação de peça",
    "fazer_os": "fechou OS em campo",
    "nao_atendido": "não conseguiu atender",
}


@relatorios_bp.route("/relatorios/atividade-recente", methods=["GET"])
def atividade_recente():
    """Feed dos últimos eventos reais -- pedido de 2026-09-04 pra Central de
    Comando: "quero estatísticas, coisas acontecendo", não só números
    parados no topo da tela. Junta desfecho de atendimento (técnico em
    campo) e orçamento aprovado (site/painel), os mais recentes primeiro.
    """
    with db_conn() as conn:
        desfechos = fetch_all(conn, sql("""
            SELECT sd.registrado_em AS quando, sd.desfecho, s.cliente
              FROM servico_desfecho sd
              JOIN servicos s ON s.id = sd.servico_id
             WHERE sd.registrado_em IS NOT NULL
             ORDER BY sd.registrado_em DESC LIMIT 15
        """))
        orcamentos = fetch_all(conn, sql("""
            SELECT os.orcamento_aprovado_em AS quando, c.nome AS cliente
              FROM ordens_servico os
              LEFT JOIN clientes c ON c.id = os.cliente_id
             WHERE os.orcamento_aprovado_em IS NOT NULL
             ORDER BY os.orcamento_aprovado_em DESC LIMIT 15
        """))

    eventos = []
    for d in desfechos:
        rotulo = _DESFECHO_ROTULO_FEED.get(d["desfecho"], d["desfecho"])
        eventos.append({
            "quando": d["quando"],
            "texto": f"{d['cliente'] or 'Cliente sem nome'} — {rotulo}",
            "tipo": d["desfecho"],
        })
    for o in orcamentos:
        eventos.append({
            "quando": o["quando"],
            "texto": f"{o['cliente'] or 'Cliente sem nome'} — orçamento aprovado",
            "tipo": "orcamento_aprovado",
        })

    eventos.sort(key=lambda e: e["quando"] or "", reverse=True)
    return jsonify({"eventos": eventos[:15]})


@relatorios_bp.route("/relatorios/mapa-calor", methods=["GET"])
def mapa_calor():
    """Pontos (lat/lng) dos atendimentos mais recentes, pra plotar como
    mancha de densidade no mapa — pedido de 2026-09-02/03: "de onde vêm
    mais chamados, por bairro/região", complementando o ranking em texto
    de /relatorios/negocio com a visão geográfica de verdade.

    `servicos` não tem coluna de data própria (só o dia da FICHA) — em vez
    de complicar com join e período, pega os N mais recentes por id, que na
    prática é "atendimentos recentes" sem depender de fichas antigas terem
    ou não `data_referencia` preenchida.
    """
    with db_conn() as conn:
        linhas = fetch_all(conn, sql("""
            SELECT lat, lng FROM servicos
             WHERE lat IS NOT NULL AND lng IS NOT NULL
             ORDER BY id DESC LIMIT 3000
        """))
    return jsonify({"pontos": linhas})


@relatorios_bp.route("/relatorios/negocio", methods=["GET"])
def relatorios_negocio():
    """Três números de negócio pedidos em 2026-09-02: previsão de
    faturamento, desempenho por setor mês a mês, e de onde vêm mais
    chamados (bairro). Junto numa rota só pra não multiplicar consulta."""
    hoje = datetime.now()
    inicio_mes = hoje.strftime("%Y-%m-01 00:00:00")

    with db_conn() as conn:
        # Previsão: valor de orçamento APROVADO este mês (ainda não é
        # faturamento realizado — é o que está pra entrar).
        previsao = fetch_one(conn, sql("""
            SELECT COALESCE(SUM(i.valor), 0) AS total, COUNT(DISTINCT os.id) AS os
              FROM ordens_servico os
              JOIN ordem_servico_itens i ON i.ordem_servico_id = os.id
             WHERE os.orcamento_aprovado_em >= ?
        """), (inicio_mes,))

        # Desempenho por setor, últimos 6 meses — mês calculado em Python
        # (mesmo motivo de sempre: SQLite e Postgres não compartilham função
        # de data), agrupado depois de trazer as linhas.
        seis_meses = (hoje - timedelta(days=180)).strftime("%Y-%m-%d 00:00:00")
        linhas_setor = fetch_all(conn, sql("""
            SELECT s.nome AS setor, sv.concluido_em
              FROM servicos sv
              JOIN setores s ON s.id = sv.setor_id
             WHERE sv.concluido_em IS NOT NULL AND sv.concluido_em >= ?
        """), (seis_meses,))

        # Ranking de bairro — de onde vêm mais chamados (cliente ligado à OS).
        bairros = fetch_all(conn, sql("""
            SELECT COALESCE(NULLIF(c.bairro, ''), NULLIF(c.cidade, ''), 'Não informado') AS local,
                   COUNT(*) AS total
              FROM ordens_servico os
              JOIN clientes c ON c.id = os.cliente_id
             GROUP BY local
             ORDER BY total DESC
             LIMIT 10
        """))

    por_mes_setor = defaultdict(lambda: defaultdict(int))
    for l in linhas_setor:
        mes = (l.get("concluido_em") or "")[:7]
        if mes:
            por_mes_setor[mes][l["setor"]] += 1
    setores_no_periodo = sorted({s for meses in por_mes_setor.values() for s in meses})
    comparativo_setores = [
        {"mes": mes, **{s: por_mes_setor[mes].get(s, 0) for s in setores_no_periodo}}
        for mes in sorted(por_mes_setor.keys())
    ]

    return jsonify({
        "previsao_faturamento": {
            "valor": round(float(previsao["total"] or 0), 2),
            "orcamentos": previsao["os"],
            "mes_referencia": hoje.strftime("%m/%Y"),
        },
        "comparativo_setores": comparativo_setores,
        "setores": setores_no_periodo,
        "ranking_bairros": bairros,
    })


@relatorios_bp.route("/pedidos-peca-os/<int:pedido_id>/pedido", methods=["POST"])
def marcar_peca_os_pedida(pedido_id):
    """Igual a marcar_peca_pedida, mas para peça pedida direto na OS (sem
    visita de técnico) — tabela pedido_peca_os em vez de servico_desfecho.
    Duplicado de propósito: juntar os dois num só exigiria um id combinado
    (tipo "t12"/"o5") em toda rota que hoje espera INTEGER puro."""
    from flask import session

    data = request.get_json(silent=True) or {}
    foto = data.get("foto")
    if foto is not None:
        foto = _foto_valida(foto)
        if not foto:
            return jsonify({"erro": "Foto inválida ou grande demais."}), 400

    with db_conn() as conn:
        dado = fetch_one(conn, sql("""
            SELECT p.id, p.peca, p.descricao, p.pedido_em,
                   c.nome AS cliente, os.tipo_aparelho, os.modelo
              FROM pedido_peca_os p
              JOIN ordens_servico os ON os.id = p.ordem_servico_id
              JOIN clientes c ON c.id = os.cliente_id
             WHERE p.id = ?
        """), (pedido_id,))

    if not dado:
        return jsonify({"erro": "Pedido de peça não encontrado"}), 404
    if dado["pedido_em"]:
        return jsonify({"erro": "Esta peça já foi marcada como pedida.",
                        "pedido_em": dado["pedido_em"]}), 409

    quem = (session.get("usuario_nome") or "").strip()[:80]
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with db_conn(commit=True) as conn:
        execute(conn, sql("UPDATE pedido_peca_os SET pedido_em = ?, "
                          "pedido_por = ?, pedido_foto = ? WHERE id = ?"),
                (agora, quem, foto, pedido_id))

    aviso = None
    try:
        from services.planilha import registrar_peca_solicitada
        r = registrar_peca_solicitada({
            "servico_id": None,
            "cliente": dado.get("cliente"),
            "peca": dado.get("peca"),
            "aparelho": " ".join(x for x in [dado.get("tipo_aparelho"),
                                             dado.get("modelo")] if x),
            "numero_os": None,
            "tecnico": None,
            "observacao": dado.get("descricao"),
            "pedido_por": quem,
        })
        if not r.get("configurada"):
            aviso = "Baixa registrada, mas a planilha não está configurada."
    except Exception as exc:
        log.exception("Falha ao gravar peça solicitada (OS) na planilha")
        aviso = f"Baixa registrada, mas não consegui escrever na planilha: {exc}"

    return jsonify({"pedido_em": agora, "pedido_por": quem, "tem_foto": bool(foto), "aviso": aviso})


@relatorios_bp.route("/desfechos/pedidos", methods=["GET"])
def listar_pedidos_de_peca():
    """Peças já marcadas como "pedidas" — vira a aba própria em Peças.

    Separada do /desfechos porque ali a foto (pedido_foto, até ~900KB em
    base64) entraria em TODA carga da aba Atendimentos mesmo quando ninguém
    pediu pra ver — aqui a lista já É a foto, então carregar junto faz
    sentido.
    """
    with db_conn() as conn:
        linhas = fetch_all(conn, sql("""
            SELECT d.servico_id, d.peca, d.observacao, d.pedido_em, d.pedido_por,
                   d.pedido_foto, s.cliente, s.endereco_completo,
                   s.tipo_aparelho, s.modelo, s.numero_os,
                   t.nome AS tecnico, t.cor AS tecnico_cor
              FROM servico_desfecho d
              JOIN servicos s ON s.id = d.servico_id
              LEFT JOIN fichas f ON f.id = s.ficha_id
              LEFT JOIN tecnicos t ON t.id = f.tecnico_id
             WHERE d.pedido_em IS NOT NULL
             ORDER BY d.pedido_em DESC
        """))
        for l in linhas:
            l["pedido_os_id"] = None
            l["chave_chegada"] = f"t{l['servico_id']}"

        # LEFT JOIN em tudo (pedido de 2026-09-02, revisão): pedido manual
        # pode não ter OS nenhuma por trás (cliente_id direto na própria
        # tabela) ou não ter cliente nenhum (reposição de estoque) — INNER
        # JOIN sumia com essas linhas inteiras da lista, silenciosamente.
        # COALESCE pega o cliente de onde tiver: direto, ou via OS antiga.
        pecas_os = fetch_all(conn, sql("""
            SELECT p.id AS pedido_os_id, p.peca, p.descricao AS observacao,
                   p.pedido_em, p.pedido_por, p.pedido_foto,
                   COALESCE(c.nome, c2.nome) AS cliente,
                   os.tipo_aparelho, os.modelo
              FROM pedido_peca_os p
              LEFT JOIN ordens_servico os ON os.id = p.ordem_servico_id
              LEFT JOIN clientes c  ON c.id = p.cliente_id
              LEFT JOIN clientes c2 ON c2.id = os.cliente_id
             WHERE p.pedido_em IS NOT NULL
        """))
        for p in pecas_os:
            p["servico_id"] = None
            p["endereco_completo"] = None
            p["numero_os"] = None
            p["tecnico"] = None
            p["tecnico_cor"] = None
            p["chave_chegada"] = f"o{p['pedido_os_id']}"
            linhas.append(p)
        linhas.sort(key=lambda l: l["pedido_em"] or "", reverse=True)

        # "Chegou?" (pedido de 2026-09-03, pra Pedidos com comprovante — a
        # mesma pergunta que /pedidos/chegada já resolve pra Panasonic, só
        # que sem nota fiscal aqui, por isso a chave sintética t<servico_id>/
        # o<pedido_os_id> em vez da chave de 44 dígitos da NF-e. Reaproveita
        # a MESMA tabela pecas_chegada e a MESMA rota de marcar/desmarcar —
        # zero schema novo.
        if linhas:
            chaves = [l["chave_chegada"] for l in linhas]
            marcadores = ",".join("?" * len(chaves))
            chegadas = fetch_all(conn, sql(
                f"SELECT chave, chegou_em FROM pecas_chegada WHERE chave IN ({marcadores})"),
                chaves)
            chegou_por_chave = {c["chave"]: c["chegou_em"] for c in chegadas}
            for l in linhas:
                l["chegou_em"] = chegou_por_chave.get(l["chave_chegada"]) or None

    return jsonify({"pedidos": linhas})
